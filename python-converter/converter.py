#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
УНИФИЦИРОВАННЫЙ КОНВЕРТЕР EXCEL/CSV В JSON
Версия: 7.1.1 (Исправлен формат телефона: +7XXXXXXXXXX)
"""
# === ИСПРАВЛЕНИЕ ДЛЯ PYINSTALLER + --windowed ===
if __name__ == '__main__':
    import sys
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        class DummyStream:
            def write(self, text): pass
            def flush(self): pass
        if sys.stderr is None:
            sys.stderr = DummyStream()
        if sys.stdout is None:
            sys.stdout = DummyStream()
# ========================================================================

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import sys
import os
from pathlib import Path
import queue
import pandas as pd
import json
import re
import argparse
from datetime import datetime, timedelta
from openpyxl import load_workbook
from collections import Counter, defaultdict
import subprocess
import csv
from typing import Any, Dict, List, Optional, Tuple, Union
import logging
import traceback
import webbrowser
import time
import random
import requests
import zipfile
import tempfile
import shutil
from packaging import version


# ============================================================================
# СПРАВОЧНИК ОСНОВАНИЙ АННУЛИРОВАНИЯ
# ============================================================================
CANCELLATION_REASONS = {
    "1": "Несоответствие лица, получившего разрешение, требованиям частей 1 и 2 статьи 49 Федерального закона N 209-ФЗ",
    "2": "Подача лицом, получившим разрешение, заявления об аннулировании разрешения",
    "3": "Ликвидация получившего разрешение юридического лица или смерти физического лица, зарегистрированного в качестве индивидуального предпринимателя, получившего разрешение"
}


def check_python_installation():
    """Проверка наличия Python и предложение установки"""
    try:
        # Проверяем версию Python
        python_version = sys.version_info
        if python_version.major < 3 or (python_version.major == 3 and python_version.minor < 7):
            print("Внимание: требуется Python 3.7 или выше")
            print(f"Текущая версия: {sys.version}")
            
            response = input("Установить Python автоматически? (y/n): ")
            if response.lower() == 'y':
                install_python()
            else:
                print("Продолжаем с текущей версией...")
    except Exception as e:
        print(f"Ошибка проверки Python: {e}")


def install_python():
    """Скачивание и установка Python"""
    try:
        print("Скачивание Python...")
        
        # URL для скачивания Python 3.9 (совместимый с Windows)
        python_url = "https://www.python.org/ftp/python/3.9.13/python-3.9.13-amd64.exe"
        installer_path = "python_installer.exe"
        
        # Скачиваем установщик
        response = requests.get(python_url, stream=True)
        with open(installer_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        print(f"Установщик сохранен: {installer_path}")
        print("Запуск установки...")
        
        # Запускаем установщик с параметрами для тихой установки
        subprocess.run([installer_path, "/quiet", "InstallAllUsers=1", "PrependPath=1"])
        
        print("Установка завершена. Перезапустите программу.")
        sys.exit(0)
        
    except Exception as e:
        print(f"Ошибка установки Python: {e}")
        print("Пожалуйста, установите Python вручную с https://python.org")


def load_config(config_file):
    """Загрузить конфигурацию из JSON файла"""
    with open(config_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def apply_config(args, config):
    """Применить конфигурацию к аргументам"""
    for key, value in config.items():
        if hasattr(args, key):
            setattr(args, key, value)


def main():
    # Проверяем Python перед запуском
    check_python_installation()
    
    # === ПЕРЕД ЗАПУСКОМ ARGPARSE: УСТАНАВЛИВАЕМ DUMMY STREAM ===
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        class DummyStream:
            def write(self, text): pass
            def flush(self): pass
        if sys.stderr is None:
            sys.stderr = DummyStream()
        if sys.stdout is None:
            sys.stdout = DummyStream()
    # ========================================================================

    # Создаем парсер
    parser = argparse.ArgumentParser(
        description='Конвертер Excel/CSV файлов в JSON формат для охотников',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('inputfile', nargs='?', help='Путь к входному файлу Excel/CSV')
    parser.add_argument('-c', '--config', help='Путь к конфигурационному файлу JSON')
    parser.add_argument('-o', '--output', help='Путь к выходной папке', default='output')
    parser.add_argument('--mode', choices=['smart'], default='smart',
                       help='Режим конвертации')
    parser.add_argument('--report', action='store_true', help='Создать отчет о конвертации')
    parser.add_argument('--sheet', help='Имя листа в Excel файле')
    parser.add_argument('--split', type=int, help='Разделить результат на указанное количество файлов')
    parser.add_argument('--postal', action='store_true', help='Включить почтовые индексы в выходные данные')
    parser.add_argument('--oktmo', action='store_true', help='Включить коды ОКТМО в выходные данные')
    parser.add_argument('--region', type=int, help='Код региона для фильтрации данных')

    # Пытаемся получить аргументы
    try:
        args, remaining_argv = parser.parse_known_args()
    except SystemExit:
        # Запускаем GUI если нет аргументов
        try:
            converter = ExcelConverterGUI()
            converter.root.mainloop()
        except Exception as e:
            print(f"Ошибка запуска GUI: {e}")
            import traceback
            traceback.print_exc()
            return
        return

    # Если config указан, загружаем его
    if args.config:
        try:
            config = load_config(args.config)
            for key, value in config.items():
                if key != 'config':
                    setattr(args, key, value)
        except Exception as e:
            print(f"Ошибка загрузки конфигурации: {e}")
            sys.exit(1)

    # Проверяем существование входного файла
    if args.inputfile and not Path(args.inputfile).exists():
        print(f"Ошибка: входной файл не найден: {args.inputfile}")
        sys.exit(1)

    # Запускаем конвертацию
    try:
        converter = ExcelConverterGUI()
        converter.run_conversion(args)
    except Exception as e:
        print(f"Критическая ошибка: {e}")
        sys.exit(1)


# ============================================================================
# КЛАСС ГРАФИЧЕСКОГО ИНТЕРФЕЙСА С ПРОКРУТКОЙ И МАСШТАБИРОВАНИЕМ
# ============================================================================
class ExcelConverterGUI:
    def __init__(self):
        self.root = None
        self.log_queue = queue.Queue()
        self.conversion_thread = None
        self.is_running = False
        self.args = None
        self.oktmo_csv_path = None
        self.selected_region = None
        self.progress_var = None
        self.progress_label = None
        self.start_time = None
        self.hunting_tickets_data = []  # Данные для huntingtickets.json
        self.enriched_data = None
        self.original_df = None
        self.nationality_df = None  # Справочник национальностей
        self.nationality_file = None  # Файл справочника национальностей
        self.scale_factor = 1.0  # Фактор масштабирования
        self.base_font_size = 10  # Базовый размер шрифта
        self.base_padding = 10  # Базовый отступ
        self.min_window_width = 1200
        self.min_window_height = 800

    def create_gui(self):
        """Создание графического интерфейса с масштабируемым дизайном"""
        self.root = tk.Tk()
        self.root.title("Конвертер Excel/CSV в JSON - Охотничьи билеты")
        self.root.geometry("1300x900")
        self.root.minsize(self.min_window_width, self.min_window_height)
        
        # Устанавливаем стиль
        style = ttk.Style()
        style.theme_use('clam')
        
        # Настраиваем цвета
        bg_color = '#F0F8FF'
        button_color = '#4A90E2'
        button_hover = '#357ABD'
        accent_color = '#FF6B35'
        
        style.configure('TFrame', background=bg_color)
        style.configure('TLabel', background=bg_color, font=('Arial', self.base_font_size))
        style.configure('TLabelframe', background=bg_color, relief=tk.GROOVE, borderwidth=2)
        style.configure('TLabelframe.Label', background=bg_color, font=('Arial', self.base_font_size, 'bold'))
        
        # Создаем главный контейнер с прокруткой и масштабированием
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Создаем Canvas для поддержки масштабирования
        self.canvas = tk.Canvas(main_container, bg=bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        # Окно для содержимого на Canvas
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)

        # Привязываем изменение размера окна
        self.root.bind('<Configure>', self.on_window_resize)

        # Меню
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Файл", menu=file_menu)
        file_menu.add_command(label="Открыть Excel/CSV", command=self.select_input_file)
        file_menu.add_command(label="Открыть ОКТМО", command=self.select_oktmo_file)
        file_menu.add_command(label="Открыть Национальность", command=self.select_nationality_file)
        file_menu.add_separator()
        file_menu.add_command(label="Сохранить обогащенный", command=self.save_enriched_file)
        file_menu.add_command(label="Сохранить лог", command=self.save_log)
        file_menu.add_separator()
        file_menu.add_command(label="Выход", command=self.root.quit)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Справка", menu=help_menu)
        help_menu.add_command(label="О программе", command=self.show_about)
        help_menu.add_command(label="Инструкция", command=self.show_instruction)
        help_menu.add_command(label="Проверить Python", command=self.check_python)

        # === МАСШТАБИРУЕМЫЙ ИНТЕРФЕЙС ===

        # Верхняя панель - файлы
        file_frame = tk.LabelFrame(self.scrollable_frame, text="Файлы",
                                   padx=15, pady=15, bg=bg_color, fg="#003366",
                                   font=("Arial", 11, "bold"))
        file_frame.pack(fill=tk.X, pady=(0, 10))

        # Excel/CSV файл
        ttk.Label(file_frame, text="Исходный файл:", font=('Arial', self.base_font_size)).grid(
            row=0, column=0, sticky=tk.W, pady=4)
        self.excel_entry = ttk.Entry(file_frame, font=('Arial', self.base_font_size))
        self.excel_entry.grid(row=0, column=1, padx=(10, 5), sticky=tk.W+tk.E, pady=4)
        self.create_styled_button(file_frame, "Выбрать...", self.select_input_file,
                                 width=15).grid(row=0, column=2, padx=(0, 10), pady=4)

        # Выходная папка
        ttk.Label(file_frame, text="Выходная папка:", font=('Arial', self.base_font_size)).grid(
            row=1, column=0, sticky=tk.W, pady=4)
        self.output_entry = ttk.Entry(file_frame, font=('Arial', self.base_font_size))
        self.output_entry.grid(row=1, column=1, padx=(10, 5), sticky=tk.W+tk.E, pady=4)
        self.output_entry.insert(0, "output")

        # ОКТМО файл
        ttk.Label(file_frame, text="Справочник ОКТМО:", font=('Arial', self.base_font_size)).grid(
            row=2, column=0, sticky=tk.W, pady=4)
        self.oktmo_entry = ttk.Entry(file_frame, font=('Arial', self.base_font_size), state='readonly')
        self.oktmo_entry.grid(row=2, column=1, padx=(10, 5), sticky=tk.W+tk.E, pady=4)
        self.create_styled_button(file_frame, "Загрузить...", self.select_oktmo_file,
                                 width=15).grid(row=2, column=2, padx=(0, 10), pady=4)

        # Файл национальностей
        ttk.Label(file_frame, text="Национальность:", font=('Arial', self.base_font_size)).grid(
            row=3, column=0, sticky=tk.W, pady=4)
        self.nationality_entry = ttk.Entry(file_frame, font=('Arial', self.base_font_size), state='readonly')
        self.nationality_entry.grid(row=3, column=1, padx=(10, 5), sticky=tk.W+tk.E, pady=4)
        self.create_styled_button(file_frame, "Загрузить...", self.select_nationality_file,
                                 width=15).grid(row=3, column=2, padx=(0, 10), pady=4)

        file_frame.grid_columnconfigure(1, weight=1)

        # Настройки конвертации
        settings_frame = tk.LabelFrame(self.scrollable_frame, text="Настройки конвертации",
                                       padx=15, pady=15, bg=bg_color, fg="#003366",
                                       font=("Arial", 11, "bold"))
        settings_frame.pack(fill=tk.X, pady=(0, 10))

        # Убраны режимы - оставлен только "Умный"
        ttk.Label(settings_frame, text="Режим конвертации:", font=('Arial', self.base_font_size)).grid(
            row=0, column=0, sticky=tk.W, pady=4)
        self.mode_label = ttk.Label(settings_frame, text="Умный", font=('Arial', self.base_font_size, 'bold'), 
                                   foreground="#4CAF50")
        self.mode_label.grid(row=0, column=1, sticky=tk.W, pady=4, padx=(10, 0))

        # Опции чекбоксов
        options_frame = ttk.Frame(settings_frame)
        options_frame.grid(row=1, column=0, columnspan=4, pady=10, sticky=tk.W)

        self.postal_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Почтовые индексы",
                        variable=self.postal_var, style='TCheckbutton').pack(side=tk.LEFT, padx=(0, 20))

        self.oktmo_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="ОКТМО коды",
                        variable=self.oktmo_var, style='TCheckbutton').pack(side=tk.LEFT, padx=(0, 20))

        self.report_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Создать отчет",
                        variable=self.report_var, style='TCheckbutton').pack(side=tk.LEFT)

        # Выбор региона - сортировка по номеру региона
        ttk.Label(settings_frame, text="Регион РФ:", font=('Arial', self.base_font_size)).grid(
            row=2, column=0, sticky=tk.W, pady=4)
        self.region_var = tk.StringVar()
        self.region_combo = ttk.Combobox(settings_frame, textvariable=self.region_var,
                                        font=('Arial', self.base_font_size), state="readonly")
        self.region_combo.grid(row=2, column=1, columnspan=2, sticky=tk.W+tk.E, pady=4, padx=(10, 0))
        
        # Сортируем регионы по номеру
        sorted_regions = [""] + [f"{code} - {name}" for code, name in sorted(RUSSIAN_REGIONS.items())]
        self.region_combo['values'] = sorted_regions
        self.region_combo.set("")

        settings_frame.grid_columnconfigure(1, weight=1)

        # Панель кнопок проверки
        check_frame = ttk.Frame(self.scrollable_frame)
        check_frame.pack(fill=tk.X, pady=(0, 10))

        self.create_styled_button(check_frame, "Проверить данные", self.check_data,
                                 width=20).pack(side=tk.LEFT, padx=(0, 10))
        self.create_styled_button(check_frame, "Проверить конвертацию", self.verify_conversion,
                                 width=20).pack(side=tk.LEFT, padx=(0, 10))

        # Прогресс бар
        progress_frame = tk.LabelFrame(self.scrollable_frame, text="Прогресс конвертации",
                                       padx=15, pady=15, bg=bg_color, fg="#003366",
                                       font=("Arial", 11, "bold"))
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var,
                                           maximum=100, style='green.Horizontal.TProgressbar')
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))

        info_frame = ttk.Frame(progress_frame)
        info_frame.pack(fill=tk.X)

        self.progress_label = ttk.Label(info_frame, text="Готов к работе",
                                        font=("Arial", self.base_font_size, "bold"), foreground="#003366")
        self.progress_label.pack(side=tk.LEFT)

        self.time_label = ttk.Label(info_frame, text="Осталось: --:--",
                                    font=("Arial", self.base_font_size), foreground="#666666")
        self.time_label.pack(side=tk.RIGHT)

        # Лог конвертации
        log_frame = tk.LabelFrame(self.scrollable_frame, text="Лог конвертации",
                                  padx=15, pady=15, bg=bg_color, fg="#003366",
                                  font=("Arial", 11, "bold"))
        log_frame.pack(fill=tk.BOTH, expand=True)

        log_scrollbar = ttk.Scrollbar(log_frame)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.text_area = tk.Text(
            log_frame,
            wrap=tk.WORD,
            yscrollcommand=log_scrollbar.set,
            font=("Consolas", 9),
            bg="white",
            fg="black",
            height=15
        )
        self.text_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        log_scrollbar.config(command=self.text_area.yview)

        # Панель основных кнопок
        button_frame = ttk.Frame(self.scrollable_frame)
        button_frame.pack(fill=tk.X, pady=(15, 0))

        # Левая группа кнопок
        left_btn_frame = ttk.Frame(button_frame)
        left_btn_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.convert_button = self.create_styled_button(left_btn_frame, "🚀 Начать конвертацию",
                                                        self.start_conversion,
                                                        width=25, height=2, bg="#4CAF50", fg="white")
        self.convert_button.pack(side=tk.LEFT, padx=(0, 10))

        self.stop_button = self.create_styled_button(left_btn_frame, "⏹ Остановить",
                                                     self.stop_conversion,
                                                     width=20, height=2, bg="#F44336", fg="white",
                                                     state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=(0, 10))

        # Правая группа кнопок
        right_btn_frame = ttk.Frame(button_frame)
        right_btn_frame.pack(side=tk.RIGHT)

        self.save_enriched_btn = self.create_styled_button(right_btn_frame, "💾 Сохранить обогащенный", 
                                 self.save_enriched_file, width=22, state='disabled')
        self.save_enriched_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.create_styled_button(right_btn_frame, "📋 Сохранить лог", 
                                 self.save_log, width=15).pack(side=tk.LEFT, padx=(0, 5))

        self.create_styled_button(right_btn_frame, "❓ О программе", 
                                 self.show_about, width=15).pack(side=tk.LEFT, padx=(0, 5))

        self.create_styled_button(right_btn_frame, "❌ Закрыть", 
                                 self.root.quit, width=15, bg="#607D8B").pack(side=tk.LEFT)

        # Статусная строка
        self.status_label = tk.Label(self.scrollable_frame, text="Готов к работе",
                                     relief=tk.SUNKEN, font=("Arial", 9),
                                     bg="#F0F8FF", fg="#003366", padx=10, pady=5)
        self.status_label.pack(fill=tk.X, pady=(10, 0))

        # Упаковка Canvas и Scrollbar
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Запускаем обработчик сообщений
        self.root.after(100, self.process_log_queue)

        # Привязываем клавиши
        self.root.bind('<Control-o>', lambda e: self.select_input_file())
        self.root.bind('<Control-s>', lambda e: self.save_log())
        self.root.bind('<Control-q>', lambda e: self.root.quit())

        # Первоначальное обновление макета
        self.update_layout()

    def on_window_resize(self, event=None):
        """Обработчик изменения размера окна"""
        if event and event.widget == self.root:
            # Получаем текущий размер окна
            window_width = self.root.winfo_width()
            window_height = self.root.winfo_height()
            
            # Рассчитываем масштабный коэффициент
            width_scale = window_width / self.min_window_width
            height_scale = window_height / self.min_window_height
            
            # Используем минимальный масштаб
            self.scale_factor = min(width_scale, height_scale)
            
            # Ограничиваем масштаб
            self.scale_factor = max(0.8, min(self.scale_factor, 1.5))
            
            # Обновляем макет
            self.update_layout()
            
            # Обновляем размер Canvas окна
            self.canvas.itemconfig(self.canvas_window, width=window_width - 50)

    def update_layout(self):
        """Обновление макета с учетом масштабирования"""
        # Обновляем ширину виджетов ввода
        for entry_widget in [self.excel_entry, self.output_entry, self.oktmo_entry, self.nationality_entry]:
            if entry_widget:
                entry_widget.config(width=int(70 * self.scale_factor))
        
        # Обновляем ширину ComboBox региона
        if self.region_combo:
            self.region_combo.config(width=int(40 * self.scale_factor))
        
        # Обновляем размер прогресс-бара
        if self.progress_bar:
            # Получаем ширину окна
            window_width = self.root.winfo_width()
            # Устанавливаем длину прогресс-бара (минус отступы)
            progress_width = max(300, int(window_width * 0.85))
            self.progress_bar.config(length=progress_width)

    def create_styled_button(self, parent, text, command, **kwargs):
        """Создание стилизованной кнопки"""
        width = kwargs.get('width', 10)
        height = kwargs.get('height', 1)
        bg = kwargs.get('bg', '#4A90E2')
        fg = kwargs.get('fg', 'white')
        state = kwargs.get('state', tk.NORMAL)
        
        # Применяем масштабирование к размеру кнопки
        scaled_width = int(width * self.scale_factor)
        scaled_height = height
        
        btn = tk.Button(parent, text=text, command=command,
                       bg=bg, fg=fg, font=('Arial', self.base_font_size, 'bold'),
                       relief=tk.RAISED, borderwidth=2,
                       padx=int(15 * self.scale_factor), 
                       pady=int(8 * self.scale_factor),
                       width=scaled_width, height=scaled_height,
                       cursor='hand2')
        
        if state == tk.DISABLED:
            btn.config(state=tk.DISABLED, bg='#CCCCCC')
        
        # Эффект при наведении
        def on_enter(e):
            if btn['state'] != tk.DISABLED:
                btn.config(bg='#357ABD')
        
        def on_leave(e):
            if btn['state'] != tk.DISABLED:
                btn.config(bg=bg)
        
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        
        return btn

    def show_about(self):
        """Показать подробную информацию о программе"""
        about_text = """📊 КОНВЕРТЕР EXCEL/CSV В JSON - ВЕРСИЯ 7.1.1

🎯 НАЗНАЧЕНИЕ ПРОГРАММЫ:
Конвертация данных из Excel/CSV в два JSON файла:
1. hunters.json - полная информация об охотниках
2. huntingtickets.json - информация о выданных билетах

📋 ОСНОВНЫЕ ВОЗМОЖНОСТИ:

📁 ФАЙЛОВЫЕ ОПЕРАЦИИ:
• Открытие Excel/CSV файлов
• Загрузка справочников (ОКТМО, национальности)
• Сохранение обогащенного файла
• Экспорт логов

⚙️ НАСТРОЙКИ КОНВЕРТАЦИИ:
• Режим работы: Умный
• Фильтрация по регионам РФ
• Автоматическое обогащение данных
• Создание детальных отчетов

🔍 ПРОВЕРКИ И ВАЛИДАЦИЯ:
• Проверка обязательных полей
• Валидация форматов дат
• Проверка СНИЛС, серий, номеров
• Контроль национальностей по справочнику

🔄 ПРОЦЕСС КОНВЕРТАЦИИ:
1. Выбор исходного файла
2. Загрузка справочников (опционально)
3. Настройка параметров
4. Проверка данных
5. Конвертация с прогресс-баром
6. Сохранение результатов

📊 ВЫХОДНЫЕ ДАННЫЕ:
• hunters.json - структурированные данные охотников
• huntingtickets.json - информация о билетах
• Отчет о конвертации
• Обогащенный исходный файл

🔧 ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ:
• Поддержка Excel (xlsx, xls) и CSV
• Автоматическое определение структуры
• Обогащение данных из справочников
• Логирование операций
• Современный масштабируемый интерфейс

📞 ПОДДЕРЖКА:
Версия: 7.1.1 | © 2024 Все права защищены
Для технической поддержки обратитесь к разработчику"""

        self._show_compact_window("📋 О программе", about_text)

    def show_instruction(self):
        """Показать инструкцию"""
        instruction_text = """📘 ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ВЕРСИИ 7.1.1

🚀 БЫСТРЫЙ СТАРТ:
1. Нажмите 'Выбрать...' для выбора Excel/CSV файла
2. При необходимости загрузите справочники
3. Настройте параметры конвертации
4. Нажмите 'Начать конвертацию'

📁 РАБОТА С ФАЙЛАМИ:

1. ИСХОДНЫЙ ФАЙЛ:
   • Поддерживаются форматы: .xlsx, .xls, .csv
   • Файл может содержать английские и русские заголовки
   • Программа автоматически определит структуру

2. СПРАВОЧНИКИ:
   • ОКТМО.csv - для обогащения муниципальных кодов
   • Национальность.xls - для проверки национальностей
   • Справочники можно загрузить в любое время

⚙️ НАСТРОЙКИ:

1. РЕЖИМ КОНВЕРТАЦИИ:
   • Умный - интеллектуальная обработка с исправлением ошибок

2. ДОПОЛНИТЕЛЬНЫЕ ОПЦИИ:
   • Почтовые индексы - поиск по справочнику ОКТМО
   • ОКТМО коды - обогащение муниципальных кодов
   • Создать отчет - детальный отчет о конвертации

3. ВЫБОР РЕГИОНА:
   • Фильтрация данных по региону РФ
   • Значение региона добавляется в JSON

🔍 ПРОВЕРКА ДАННЫХ:

1. ПРЕДВАРИТЕЛЬНАЯ ПРОВЕРКА:
   • Нажмите 'Проверить данные' для анализа файла
   • Программа покажет ошибки и предупреждения

2. ОБЯЗАТЕЛЬНЫЕ ПОЛЯ:
   • Фамилия, Имя, Отчество
   • Дата рождения
   • Место рождения
   • Дата выдачи билета
   • Серия и номер билета

3. ВАЛИДАЦИЯ ФОРМАТОВ:
   • Даты: строго ГГГГ-ММ-ДД
   • СНИЛС: XXX-XXX-XXX XX
   • Телефон: +7XXXXXXXXXX (без пробелов и скобок)
   • Серии/номера: проверка длины и символов

🔄 ПРОЦЕСС КОНВЕРТАЦИИ:

1. ЗАГРУЗКА:
   • Чтение исходного файла
   • Определение кодировки (для CSV)
   • Автоматическое определение заголовков

2. ОБРАБОТКА:
   • Валидация каждой строки
   • Обогащение данных из справочников
   • Форматирование по эталону JSON

3. СОХРАНЕНИЕ:
   • hunters.json - данные охотников
   • huntingtickets.json - данные билетов
   • Отчет (если включено)
   • Обогащенный файл (по необходимости)

📊 ВЫХОДНЫЕ ФАЙЛЫ:

1. HUNTERS.JSON:
   • Полная информация об охотниках
   • Структура соответствует эталону
   • Все значения в кавычках (включая числовые)

2. HUNTINGTICKETS.JSON:
   • Информация о выданных билетах
   • Структура соответствует эталону
   • Все значения в кавычках (включая булевы)

⚠️ ВАЖНЫЕ ЗАМЕЧАНИЯ:

1. ПУСТЫЕ СТРОКИ:
   • Строки без обязательных полей пропускаются
   • После двух пустых строк обработка останавливается

2. ОБРАБОТКА ОШИБОК:
   • Все ошибки записываются в лог
   • Создается детальный отчет
   • Программа продолжает работу при некритических ошибках

3. МАСШТАБИРУЕМОСТЬ:
   • Интерфейс адаптируется под размер окна
   • Все элементы правильно масштабируются
   • Поддержка разных разрешений экрана

🆘 ПОЛУЧЕНИЕ ПОМОЩИ:
• Нажмите 'О программе' для подробной информации
• Проверьте лог конвертации для анализа ошибок
• Сохраняйте отчеты для отладки

📞 КОНТАКТЫ:
Для вопросов и предложений обращайтесь к разработчику"""

        self._show_compact_window("📘 Инструкция", instruction_text)

    def check_python(self):
        """Проверка установки Python"""
        python_version = sys.version_info
        python_path = sys.executable
        
        message = f"""🐍 ПРОВЕРКА УСТАНОВКИ PYTHON

Текущая версия: Python {python_version.major}.{python_version.minor}.{python_version.micro}
Путь к Python: {python_path}

Минимальная требуемая версия: Python 3.7

Состояние: {"✅ Установлена корректная версия" if python_version.major == 3 and python_version.minor >= 7 else "⚠️ Требуется обновление"}

Для установки/обновления Python:
1. Перейдите на https://python.org
2. Скачайте последнюю версию Python 3.x
3. Установите с опцией "Add Python to PATH"

Рекомендуемый путь установки: C:\\Python39"""

        messagebox.showinfo("Проверка Python", message)

    def _show_compact_window(self, title, text):
        """Универсальный метод для создания компактных окон"""
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("800x600")
        window.configure(bg='#F0F8FF')
        window.minsize(600, 400)

        # Добавляем иконку
        try:
            window.iconbitmap(default='icon.ico')
        except:
            pass

        text_widget = scrolledtext.ScrolledText(window, wrap=tk.WORD,
                                               font=("Arial", 10), 
                                               bg='white', fg='#333333')
        text_widget.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        text_widget.insert(tk.END, text)
        text_widget.config(state=tk.DISABLED)

        btn = self.create_styled_button(window, "Закрыть", window.destroy, width=20)
        btn.pack(pady=10)

    def select_input_file(self):
        """Выбор входного файла (Excel или CSV)"""
        filename = filedialog.askopenfilename(
            title="Выберите Excel или CSV файл",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, filename)
            # Предлагаем имя для выходной папки
            input_path = Path(filename)
            output_path = input_path.parent / f"{input_path.stem}_converted"
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, str(output_path))
            self.log_message(f"Выбран файл: {filename}", "INFO")

    def select_oktmo_file(self):
        """Выбор файла справочника ОКТМО"""
        filename = filedialog.askopenfilename(
            title="Выберите файл справочника ОКТМО",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.oktmo_csv_path = Path(filename)
            self.oktmo_entry.config(state='normal')
            self.oktmo_entry.delete(0, tk.END)
            self.oktmo_entry.insert(0, str(self.oktmo_csv_path))
            self.oktmo_entry.config(state='readonly')
            self.log_message(f"Выбран файл ОКТМО: {self.oktmo_csv_path}", "INFO")

    def select_nationality_file(self):
        """Выбор файла справочника национальностей"""
        filename = filedialog.askopenfilename(
            title="Выберите файл справочника национальностей",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.nationality_file = Path(filename)
            self.nationality_entry.config(state='normal')
            self.nationality_entry.delete(0, tk.END)
            self.nationality_entry.insert(0, str(self.nationality_file))
            self.nationality_entry.config(state='readonly')
            self.log_message(f"Выбран файл национальностей: {self.nationality_file}", "INFO")
            
            # Загружаем справочник национальностей
            try:
                if self.nationality_file.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                    self.nationality_df = pd.read_excel(self.nationality_file)
                else:
                    self.nationality_df = pd.read_csv(self.nationality_file, encoding='utf-8')
                
                self.log_message(f"Загружено национальностей: {len(self.nationality_df)}", "SUCCESS")
            except Exception as e:
                self.log_message(f"Ошибка загрузки национальностей: {str(e)}", "ERROR")
                self.nationality_df = None

    def save_enriched_file(self):
        """Сохранение обогащенного исходного файла"""
        if self.enriched_data is None:
            messagebox.showwarning("Предупреждение", "Нет обогащенных данных для сохранения")
            return

        try:
            # Определяем расширение исходного файла
            input_file = Path(self.excel_entry.get())
            if not input_file or str(input_file) == '.':
                messagebox.showwarning("Предупреждение", "Не выбран исходный файл")
                return
                
            default_ext = input_file.suffix if input_file.suffix else '.xlsx'

            filename = filedialog.asksaveasfilename(
                defaultextension=default_ext,
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("Excel 97-2003 files", "*.xls"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ],
                initialfile=f"enriched_{input_file.stem}{default_ext}"
            )

            if filename:
                output_path = Path(filename)
                if output_path.suffix.lower() == '.csv':
                    self.enriched_data.to_csv(output_path, index=False, encoding='utf-8-sig')
                else:
                    self.enriched_data.to_excel(output_path, index=False)

                messagebox.showinfo("Успех", f"✅ Обогащенный файл сохранен:\n{filename}")
                self.log_message(f"Обогащенный файл сохранен: {filename}", "SUCCESS")

        except Exception as e:
            messagebox.showerror("Ошибка", f"❌ Не удалось сохранить файл: {e}")
            self.log_message(f"Ошибка сохранения обогащенного файла: {e}", "ERROR")

    def verify_conversion(self):
        """Проверка результатов конвертации"""
        try:
            output_folder = Path(self.output_entry.get())
            hunters_file = output_folder / "hunters.json"
            tickets_file = output_folder / "huntingtickets.json"

            if not hunters_file.exists():
                self.log_message("Файл hunters.json не найден", "WARNING")
                return

            with open(hunters_file, 'r', encoding='utf-8') as f:
                hunters_data = json.load(f)

            self.log_message("="*60, "INFO")
            self.log_message("✅ ПРОВЕРКА РЕЗУЛЬТАТОВ КОНВЕРТАЦИИ", "INFO")
            self.log_message(f"Файл hunters.json: {len(hunters_data)} записей", "INFO")

            # Проверка структуры hunters.json
            if hunters_data:
                sample = hunters_data[0]
                required_hunters_fields = [
                    'surname', 'hunter_name', 'patronymic', 'birth_date',
                    'birth_place', 'date_issue_ticket', 'series_ticket', 'number_ticket'
                ]

                missing_in_json = []
                for field in required_hunters_fields:
                    if field not in sample:
                        missing_in_json.append(field)

                if missing_in_json:
                    self.log_message(f"❌ В hunters.json отсутствуют поля: {', '.join(missing_in_json)}", "ERROR")
                else:
                    self.log_message("✅ Все обязательные поля присутствуют в hunters.json", "SUCCESS")

                # Проверка формата данных
                self.log_message("🔍 Проверка форматов данных:", "INFO")
                for i, record in enumerate(hunters_data[:3]):  # Проверяем первые 3 записи
                    self.log_message(f"Запись {i+1}:", "INFO")
                    for field in ['birth_date', 'date_issue_ticket']:
                        if field in record:
                            value = record[field]
                            if value and not re.match(r'^\d{4}-\d{2}-\d{2}$', str(value)):
                                self.log_message(f"  ❌ {field}: неверный формат даты - {value}", "ERROR")
                            else:
                                self.log_message(f"  ✅ {field}: корректный формат", "SUCCESS")
                    
                    # Проверка формата телефона
                    if 'phone' in record and record['phone']:
                        phone = record['phone']
                        if not re.match(r'^\+7\d{10}$', str(phone)):
                            self.log_message(f"  ❌ phone: неверный формат телефона - {phone}", "ERROR")
                        else:
                            self.log_message(f"  ✅ phone: корректный формат (+7XXXXXXXXXX)", "SUCCESS")

            # Проверка huntingtickets.json
            if tickets_file.exists():
                with open(tickets_file, 'r', encoding='utf-8') as f:
                    tickets_data = json.load(f)
                self.log_message(f"Файл huntingtickets.json: {len(tickets_data)} записей", "INFO")

                if tickets_data:
                    sample_ticket = tickets_data[0]
                    required_ticket_fields = ['date_entry', 'series', 'number', 'date_issue']

                    missing_in_tickets = []
                    for field in required_ticket_fields:
                        if field not in sample_ticket:
                            missing_in_tickets.append(field)

                    if missing_in_tickets:
                        self.log_message(f"❌ В huntingtickets.json отсутствуют поля: {', '.join(missing_in_tickets)}", "ERROR")
                    else:
                        self.log_message("✅ Все обязательные поля присутствуют в huntingtickets.json", "SUCCESS")
                    
                    # Проверка что все значения в кавычках (включая булевы)
                    if 'is_belonged_to_indigenous_people' in sample_ticket:
                        value = sample_ticket['is_belonged_to_indigenous_people']
                        if isinstance(value, str) and value.lower() in ['true', 'false']:
                            self.log_message("✅ Булево значение коренных народов в кавычках", "SUCCESS")
            else:
                self.log_message("⚠️ Файл huntingtickets.json не создан", "WARNING")

            self.log_message("="*60, "INFO")

        except Exception as e:
            self.log_message(f"❌ Ошибка при проверке конвертации: {str(e)}", "ERROR")

    def check_data(self):
        """Проверка данных перед конвертацией"""
        if not self.excel_entry.get():
            messagebox.showwarning("Предупреждение", "Выберите файл для проверки")
            return

        try:
            input_file = Path(self.excel_entry.get())
            if not input_file.exists():
                messagebox.showerror("Ошибка", f"Файл не найден: {input_file}")
                return

            self.log_message("="*60, "INFO")
            self.log_message("🔍 НАЧАЛО ПРОВЕРКИ ДАННЫХ", "INFO")
            self.log_message(f"Файл: {input_file}", "INFO")
            self.log_message("="*60, "INFO")

            # Загружаем данные
            df = self._load_data_for_check(input_file)
            self.original_df = df.copy()

            if df.empty:
                self.log_message("❌ Файл пуст или содержит некорректные данные", "ERROR")
                return

            # Проверяем обязательные поля
            required_fields = [
                'surname', 'hunter_name', 'patronymic',
                'birth_date', 'birth_place',
                'date_issue_ticket', 'series_ticket', 'number_ticket'
            ]

            missing_fields = []
            for field in required_fields:
                if field not in df.columns:
                    missing_fields.append(field)

            if missing_fields:
                self.log_message(f"⚠️ Отсутствуют поля в файле: {', '.join(missing_fields)}", "WARNING")
            else:
                self.log_message("✅ Все основные поля присутствуют", "SUCCESS")

                # Проверяем заполненность полей
                empty_rows = 0
                total_checked = 0
                mandatory_errors = []
                
                for idx, row in df.iterrows():
                    row_num = idx + 2  # Excel строки с 1 + заголовок
                    total_checked += 1

                    # Проверяем две пустые строки подряд
                    all_empty = True
                    for field in df.columns:
                        if field in row:
                            value = row[field]
                            if not pd.isna(value) and str(value).strip() != '':
                                all_empty = False
                                break

                    if all_empty:
                        empty_rows += 1
                        if empty_rows >= 2:
                            self.log_message(f"⚠️ Обнаружены 2 пустые строки подряд (строка {row_num}). Проверка остановлена.", "WARNING")
                            break
                    else:
                        empty_rows = 0

                    # Проверяем обязательные поля
                    missing_values = []
                    for field in required_fields:
                        if field in df.columns:
                            value = row.get(field)
                            if pd.isna(value) or str(value).strip() == '':
                                missing_values.append(field)
                                mandatory_errors.append((row_num, field))

                    if missing_values:
                        self.log_message(f"⚠️ Строка {row_num}: отсутствуют значения в полях: {', '.join(missing_values)}", "WARNING")

                # Вывод статистики по обязательным полям
                if mandatory_errors:
                    self.log_message(f"⚠️ Всего ошибок в обязательных полях: {len(mandatory_errors)}", "WARNING")
                    # Группируем по строкам
                    error_by_row = defaultdict(list)
                    for row_num, field in mandatory_errors:
                        error_by_row[row_num].append(field)
                    
                    for row_num, fields in list(error_by_row.items())[:10]:  # Показываем первые 10
                        self.log_message(f"  Строка {row_num}: {', '.join(fields)}", "INFO")

            self.log_message(f"📊 Всего строк: {len(df)}", "INFO")
            self.log_message(f"📊 Проверено строк: {total_checked}", "INFO")
            self.log_message("="*60, "INFO")
            self.log_message("✅ ПРОВЕРКА ЗАВЕРШЕНА", "INFO")

        except Exception as e:
            self.log_message(f"❌ Ошибка при проверке: {str(e)}", "ERROR")
            self.log_message(traceback.format_exc(), "ERROR")

    def _load_data_for_check(self, input_file: Path) -> pd.DataFrame:
        """Загрузка данных для проверки (Excel или CSV)"""
        try:
            file_extension = input_file.suffix.lower()

            if file_extension in ['.xlsx', '.xlsm', '.xls']:
                return self._load_excel_for_check(input_file)
            elif file_extension == '.csv':
                return self._load_csv_for_check(input_file)
            else:
                self.log_message(f"❌ Неподдерживаемый формат файла: {file_extension}", "ERROR")
                return pd.DataFrame()

        except Exception as e:
            self.log_message(f"❌ Ошибка загрузки файла: {str(e)}", "ERROR")
            return pd.DataFrame()

    def _load_excel_for_check(self, input_file: Path) -> pd.DataFrame:
        """Загрузка Excel для проверки"""
        try:
            workbook = load_workbook(input_file, read_only=True, data_only=True)
            sheet = workbook.active
            data = list(sheet.values)

            if not data:
                return pd.DataFrame()

            # Проверяем структуру с двумя строками заголовков
            if len(data) >= 2:
                first_row = [str(cell) if cell is not None else "" for cell in data[0]]
                second_row = [str(cell) if cell is not None else "" for cell in data[1]]

                # Если первая строка содержит латиницу, а вторая - кириллицу
                latin_count = sum(1 for cell in first_row if any('a' <= char.lower() <= 'z' for char in cell))
                cyrillic_count = sum(1 for cell in second_row if any('\u0400' <= char <= '\u04FF' for char in cell))

                if latin_count > 0 and cyrillic_count > 0:
                    # Используем первую строку как заголовки (английские названия)
                    self.log_message("Обнаружены две строки заголовков. Использую первую строку (английскую) как ключи JSON.", "INFO")
                    columns = first_row
                    data_rows = data[2:]  # Данные начинаются с третьей строки
                else:
                    # Используем первую строку как заголовки
                    columns = first_row
                    data_rows = data[1:]
            else:
                # Только одна строка заголовков
                columns = [str(cell) if cell is not None else "" for cell in data[0]]
                data_rows = data[1:]

            df = pd.DataFrame(data_rows, columns=columns)
            workbook.close()
            
            # Очищаем названия столбцов
            df.columns = df.columns.astype(str).str.strip()
            return df

        except Exception as e:
            self.log_message(f"❌ Ошибка загрузки Excel: {str(e)}", "ERROR")
            return pd.DataFrame()

    def _load_csv_for_check(self, input_file: Path) -> pd.DataFrame:
        """Загрузка CSV для проверки"""
        try:
            # Пробуем разные кодировки
            encodings = ['utf-8-sig', 'cp1251', 'windows-1251', 'utf-8']

            for encoding in encodings:
                try:
                    with open(input_file, 'r', encoding=encoding) as f:
                        sample = f.read(4096)
                        f.seek(0)

                        # Определяем разделитель
                        delimiter = ',' if ',' in sample else ';'

                        # Читаем CSV
                        df = pd.read_csv(
                            input_file,
                            delimiter=delimiter,
                            encoding=encoding,
                            on_bad_lines='skip'
                        )

                    # Если успешно прочитали, выходим из цикла
                    break
                except Exception as e:
                    if encoding == encodings[-1]:  # Если это последняя кодировка
                        raise e
                    continue

            # Очистка названий столбцов
            df.columns = df.columns.astype(str).str.strip()
            return df

        except Exception as e:
            self.log_message(f"❌ Ошибка загрузки CSV: {str(e)}", "ERROR")
            return pd.DataFrame()

    def log_message(self, message, level="INFO"):
        """Добавление сообщения в лог"""
        colors = {
            "ERROR": "red",
            "WARNING": "orange",
            "INFO": "black",
            "SUCCESS": "green"
        }
        color = colors.get(level, "black")
        tag = f"tag_{level}"

        self.text_area.tag_configure(tag, foreground=color)
        timestamp = datetime.now().strftime("%H:%M:%S")
        
        # Добавляем эмодзи для разных уровней
        emoji = {
            "ERROR": "❌ ",
            "WARNING": "⚠️ ",
            "SUCCESS": "✅ ",
            "INFO": "ℹ️ "
        }.get(level, "")
        
        formatted_msg = f"[{timestamp}] {emoji}{message}\n"

        self.text_area.insert(tk.END, formatted_msg, tag)
        self.text_area.see(tk.END)

        if self.root:
            self.root.update_idletasks()

    def update_progress(self, percent, message, time_remaining="--:--"):
        """Обновление прогресс-бара с временем"""
        if self.progress_var:
            self.progress_var.set(percent)
        if self.progress_label:
            self.progress_label.config(text=message)
        if self.time_label:
            self.time_label.config(text=f"⏱️ Осталось: {time_remaining}")
        if self.root:
            self.root.update_idletasks()

    def process_log_queue(self):
        """Обработка сообщений из очереди"""
        try:
            while True:
                message, level = self.log_queue.get_nowait()
                self.log_message(message, level)
        except queue.Empty:
            pass
        finally:
            if self.root:
                self.root.after(100, self.process_log_queue)

    def save_log(self):
        """Сохранение лога в файл"""
        try:
            log_content = self.text_area.get(1.0, tk.END)
            if not log_content.strip():
                messagebox.showwarning("Предупреждение", "Лог пуст")
                return

            filename = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")],
                initialfile="conversion_log.txt"
            )
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(log_content)
                messagebox.showinfo("Успех", f"✅ Лог сохранен в {filename}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"❌ Не удалось сохранить лог: {e}")

    def clear_text(self):
        """Очистка текстовой области"""
        self.text_area.delete(1.0, tk.END)

    def stop_conversion(self):
        """Остановка конвертации"""
        if self.is_running and self.conversion_thread:
            self.is_running = False
            self.status_label.config(text="Останавливается...", fg="orange")
            self.stop_button.config(state=tk.DISABLED)

    def start_conversion(self):
        """Запуск конвертации из GUI"""
        if not self.excel_entry.get():
            messagebox.showwarning("Предупреждение", "Выберите файл")
            return

        # Создаем объект аргументов
        import types
        self.args = types.SimpleNamespace()
        self.args.inputfile = self.excel_entry.get()
        self.args.output = self.output_entry.get()
        self.args.mode = "smart"  # Только умный режим
        self.args.report = self.report_var.get()
        self.args.sheet = None
        self.args.split = None
        self.args.postal = self.postal_var.get()
        self.args.oktmo = self.oktmo_var.get()

        # Получаем выбранный регион
        region_selection = self.region_var.get()
        if region_selection:
            try:
                region_code = region_selection.split(" - ")[0].strip()
                region_name = region_selection.split(" - ")[1].strip()
                
                if region_code in RUSSIAN_REGIONS:
                    self.args.region = int(region_code)
                    self.selected_region = region_name
                else:
                    messagebox.showerror("Ошибка", "Неверный регион")
                    return
            except:
                messagebox.showerror("Ошибка", "Неверный формат региона")
                return
        else:
            self.args.region = None
            self.selected_region = None

        # Передаем путь к ОКТМО
        if self.oktmo_csv_path:
            self.args.oktmo_csv = self.oktmo_csv_path
        else:
            self.args.oktmo_csv = None

        # Передаем путь к национальностям
        if self.nationality_file:
            self.args.nationality_file = self.nationality_file
        else:
            self.args.nationality_file = None

        # Запускаем конвертацию
        self.conversion_thread = threading.Thread(target=self._run_conversion_thread)
        self.conversion_thread.daemon = True
        self.conversion_thread.start()

        # Обновляем интерфейс
        self.convert_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.is_running = True
        self.start_time = time.time()
        self.status_label.config(text="Выполняется конвертация...", fg="blue")

    def run_conversion(self, args):
        """Запуск конвертации (для CLI)"""
        self.args = args
        self.create_gui()
        self.conversion_thread = threading.Thread(target=self._run_conversion_thread)
        self.conversion_thread.daemon = True
        self.conversion_thread.start()
        self.stop_button.config(state=tk.NORMAL)
        self.is_running = True
        self.start_time = time.time()
        self.status_label.config(text="Выполняется конвертация...", fg="blue")
        self.root.mainloop()

    def _run_conversion_thread(self):
        """Основной поток конвертации"""
        try:
            converter = ExcelToJsonConverter()
            converter.set_progress_callback(self.update_progress)

            # Получаем путь к ОКТМО
            oktmo_path = None
            if hasattr(self.args, 'oktmo_csv') and self.args.oktmo_csv:
                oktmo_path = Path(str(self.args.oktmo_csv))
            elif self.oktmo_csv_path:
                oktmo_path = self.oktmo_csv_path

            # Получаем путь к национальностям
            nationality_path = None
            if hasattr(self.args, 'nationality_file') and self.args.nationality_file:
                nationality_path = Path(str(self.args.nationality_file))
            elif self.nationality_file:
                nationality_path = self.nationality_file

            if not self.args.inputfile:
                self.log_queue.put(("Ошибка: не указан входной файл", "ERROR"))
                return

            input_file = Path(str(self.args.inputfile))
            output_folder = Path(str(self.args.output))

            result = converter.convert(
                input_file=input_file,
                output_folder=output_folder,
                sheet_name=self.args.sheet,
                mode=self.args.mode,
                create_report=self.args.report,
                split_count=self.args.split,
                include_postal=self.args.postal,
                include_oktmo=self.args.oktmo,
                region_code=self.args.region,
                oktmo_csv_path=oktmo_path,
                nationality_file=nationality_path,
                gui_callback=self.log_queue,
                selected_region=self.selected_region
            )

            if result.get('success'):
                self.log_queue.put(("✅ Конвертация успешно завершена!", "SUCCESS"))
                self.log_queue.put((f"📁 Выходная папка: {result['output_folder']}", "INFO"))
                if result.get('report_file'):
                    self.log_queue.put((f"📊 Отчет создан: {result['report_file']}", "INFO"))

                # Сохраняем обогащенные данные
                self.enriched_data = result.get('enriched_data')
                if self.enriched_data is not None:
                    self.log_queue.put(("💾 Данные обогащены. Можно сохранить обогащенный файл.", "INFO"))
                    if self.root:
                        self.root.after(100, lambda: self.save_enriched_btn.config(state='normal'))

                self.status_label.config(text="✅ Завершено успешно", fg="green")
                self.update_progress(100, "✅ Завершено успешно", "00:00")
            else:
                self.log_queue.put(("❌ Конвертация завершилась с ошибками", "ERROR"))
                self.status_label.config(text="❌ Завершено с ошибками", fg="red")
                self.update_progress(0, "❌ Завершено с ошибками", "--:--")

        except Exception as e:
            error_msg = f"❌ Критическая ошибка: {str(e)}\n{traceback.format_exc()}"
            self.log_queue.put((error_msg, "ERROR"))
            self.status_label.config(text="❌ Критическая ошибка", fg="red")
            self.update_progress(0, "❌ Критическая ошибка", "--:--")
        finally:
            self.is_running = False
            if self.root:
                self.root.after(100, lambda: self.stop_button.config(state=tk.DISABLED))
                self.root.after(100, lambda: self.convert_button.config(state=tk.NORMAL))
                if self.start_time:
                    elapsed = time.time() - self.start_time
                    minutes = int(elapsed // 60)
                    seconds = int(elapsed % 60)
                    self.log_queue.put((f"⏱️ Общее время выполнения: {minutes} мин. {seconds} сек.", "INFO"))


# ============================================================================
# НАСТРОЙКА ЛОГГИРОВАНИЯ
# ============================================================================
def setup_logging(input_file_path: Path):
    """Настройка логирования рядом с исходным файлом"""
    log_dir = input_file_path.parent
    log_file = log_dir / f"{input_file_path.stem}_conversion_log.txt"

    logger = logging.getLogger('ExcelConverter')
    logger.setLevel(logging.INFO)

    if logger.handlers:
        logger.handlers.clear()

    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_formatter = logging.Formatter('%(levelname)s: %(message)s')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(console_handler)

    return logger, log_file


# ============================================================================
# СПРАВОЧНИК РЕГИОНОВ РФ (ОБНОВЛЕННЫЙ)
# ============================================================================
RUSSIAN_REGIONS = {
    "01": "Республика Адыгея (Адыгея)",
    "02": "Республика Башкортостан",
    "03": "Республика Бурятия",
    "04": "Республика Алтай",
    "05": "Республика Дагестан",
    "06": "Республика Ингушетия",
    "07": "Кабардино-Балкарская Республика",
    "08": "Республика Калмыкия",
    "09": "Карачаево-Черкесская Республика",
    "10": "Республика Карелия",
    "11": "Республика Коми",
    "12": "Республика Марий Эл",
    "13": "Республика Мордовия",
    "14": "Республика Саха (Якутия)",
    "15": "Республика Северная Осетия - Алания",
    "16": "Республика Татарстан (Татарстан)",
    "17": "Республика Тыва",
    "18": "Удмуртская Республика",
    "19": "Республика Хакасия",
    "20": "Чеченская Республика",
    "21": "Чувашская Республика - Чувашия",
    "22": "Алтайский край",
    "23": "Краснодарский край",
    "24": "Красноярский край",
    "25": "Приморский край",
    "26": "Ставропольский край",
    "27": "Хабаровский край",
    "28": "Амурская область",
    "29": "Архангельская область",
    "30": "Астраханская область",
    "31": "Белгородская область",
    "32": "Брянская область",
    "33": "Владимирская область",
    "34": "Волгоградская область",
    "35": "Вологодская область",
    "36": "Воронежская область",
    "37": "Ивановская область",
    "38": "Иркутская область",
    "39": "Калининградская область",
    "40": "Калужская область",
    "41": "Камчатский край",
    "42": "Кемеровская область - Кузбасс",
    "43": "Кировская область",
    "44": "Костромская область",
    "45": "Курганская область",
    "46": "Курская область",
    "47": "Ленинградская область",
    "48": "Липецкая область",
    "49": "Магаданская область",
    "50": "Московская область",
    "51": "Мурманская область",
    "52": "Нижегородская область",
    "53": "Новгородская область",
    "54": "Новосибирская область",
    "55": "Омская область",
    "56": "Оренбургская область",
    "57": "Орловская область",
    "58": "Пензенская область",
    "59": "Пермский край",
    "60": "Псковская область",
    "61": "Ростовская область",
    "62": "Рязанская область",
    "63": "Самарская область",
    "64": "Саратовская область",
    "65": "Сахалинская область",
    "66": "Свердловская область",
    "67": "Смоленская область",
    "68": "Тамбовская область",
    "69": "Тверская область",
    "70": "Томская область",
    "71": "Тульская область",
    "72": "Тюменская область",
    "73": "Ульяновская область",
    "74": "Челябинская область",
    "75": "Забайкальский край",
    "76": "Ярославская область",
    "77": "г. Москва",
    "78": "г. Санкт-Петербург",
    "79": "Еврейская автономная область",
    "83": "Ненецкий автономный округ",
    "86": "Ханты-Мансийский автономный округ - Югра",
    "87": "Чукотский автономный округ",
    "89": "Ямало-Ненецкий автономный округ",
    "90": "Запорожская область",
    "91": "Республика Крым",
    "92": "г. Севастополь",
    "93": "Донецкая Народная Республика",
    "94": "Луганская Народная Республика",
    "95": "Херсонская область"
}

# ============================================================================
# ВСТРОЕННЫЙ БАЗОВЫЙ СПРАВОЧНИК ОКТМО
# ============================================================================
OKTMO_BASE_DATA = """code,name,postal_codes
45000000,г Москва,101000-129999
45000001,Москва (город),101000-129999
45000002,Москва (городское поселение),101000-129999
89000000,Тюменская область,625000-627999
89010000,г Тюмень,625000
89020000,Тобольский район,626100
89030000,Ялуторовский район,627010
89620000,Саратовская область,410000-413999
89621000,Саратовский район,410000
89621168,Уметское,410038
89621169,Саратовское,410000
89621170,Вольское,412900
89700000,Волгоградская область,400000-404999
89710000,г Волгоград,400000
89720000,Волжский район,404100
"""

# Обязательные поля
REQUIRED_FIELDS = [
    'surname',          # Фамилия
    'hunter_name',      # Имя
    'patronymic',       # Отчество
    'birth_date',       # Дата рождения
    'birth_place',      # Место рождения
    'date_issue_ticket',# Дата выдачи охотничьего билета
    'series_ticket',    # Серия охотничьего билета
    'number_ticket',    # Номер охотничьего билета
]

DEFAULT_SHEET_NAME = 'Таблица'


# ============================================================================
# МЕНЕДЖЕР СПРАВОЧНИКОВ ОКТМО
# ============================================================================
class OktmoManager:
    """Менеджер для работы со справочниками ОКТМО"""
    def __init__(self, logger=None):
        self.logger = logger
        self.oktmo_df = None
        self.postal_indexes = {}
        self.load_base_data()

    def load_base_data(self):
        """Загрузка базовых данных ОКТМО"""
        try:
            from io import StringIO
            self.oktmo_df = pd.read_csv(StringIO(OKTMO_BASE_DATA), dtype={'code': str})
            self.oktmo_df['code'] = self.oktmo_df['code'].astype(str).str.strip()
            self.oktmo_df.set_index('code', inplace=True)
            self._build_postal_index()
            if self.logger:
                self.logger.info(f"Загружен базовый справочник ОКТМО: {len(self.oktmo_df)} записей")
        except Exception as e:
            if self.logger:
                self.logger.error(f"Ошибка загрузки базового справочника ОКТМО: {e}")
            self.oktmo_df = pd.DataFrame(columns=['code', 'name', 'postal_codes'])

    def load_external_oktmo(self, file_path: Path):
        """Загрузка внешнего справочника ОКТМО из CSV файла"""
        try:
            if not file_path.exists():
                if self.logger:
                    self.logger.warning(f"Файл справочника ОКТМО не найден: {file_path}")
                return False

            # Пробуем разные кодировки
            encodings = ['utf-8-sig', 'cp1251', 'windows-1251', 'utf-8']

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        sample = f.read(4096)
                        f.seek(0)

                        # Определяем разделитель
                        delimiter = ',' if ',' in sample else ';'

                        # Пробуем прочитать с этой кодировкой
                        external_df = pd.read_csv(
                            file_path,
                            dtype={'code': str},
                            delimiter=delimiter,
                            encoding=encoding,
                            on_bad_lines='skip'
                        )

                    # Если успешно прочитали, выходим из цикла
                    break
                except Exception as e:
                    if encoding == encodings[-1]:  # Если это последняя кодировка
                        raise e
                    continue

            # Переименовываем колонки
            column_mapping = {}
            for col in external_df.columns:
                col_lower = str(col).lower().strip()
                if 'код' in col_lower or 'code' in col_lower:
                    column_mapping[col] = 'code'
                elif 'наимен' in col_lower or 'name' in col_lower:
                    column_mapping[col] = 'name'
                elif 'индекс' in col_lower or 'postal' in col_lower or 'почт' in col_lower:
                    column_mapping[col] = 'postal_codes'

            if column_mapping:
                external_df = external_df.rename(columns=column_mapping)

            # Проверяем наличие необходимых колонок
            if 'code' not in external_df.columns:
                external_df = external_df.rename(columns={external_df.columns[0]: 'code'})

            if 'name' not in external_df.columns:
                if len(external_df.columns) > 1:
                    external_df = external_df.rename(columns={external_df.columns[1]: 'name'})
                else:
                    external_df['name'] = ''

            if 'postal_codes' not in external_df.columns:
                external_df['postal_codes'] = ''

            # Оставляем только нужные колонки
            keep_cols = ['code', 'name', 'postal_codes']
            external_df = external_df[[col for col in keep_cols if col in external_df.columns]]

            # Очищаем данные
            external_df['code'] = external_df['code'].astype(str).str.strip()
            external_df['name'] = external_df['name'].astype(str).str.strip()

            # Удаляем пустые строки
            external_df = external_df[external_df['code'].str.len() > 0]

            # Объединяем с базовыми данными
            if self.oktmo_df is None or self.oktmo_df.empty:
                self.oktmo_df = external_df
            else:
                # Объединяем и удаляем дубликаты
                combined = pd.concat([self.oktmo_df.reset_index(), external_df])
                combined = combined.drop_duplicates(subset=['code'], keep='last')
                self.oktmo_df = combined.set_index('code')

            self._build_postal_index()

            if self.logger:
                self.logger.info(f"Загружен внешний справочник ОКТМО: {len(external_df)} записей")
                self.logger.info(f"Всего записей в справочнике: {len(self.oktmo_df)}")

            return True
        except Exception as e:
            if self.logger:
                self.logger.error(f"Ошибка загрузки внешнего справочника ОКТМО: {e}", exc_info=True)
            return False

    def _build_postal_index(self):
        """Построение словаря для поиска почтовых индексов"""
        self.postal_indexes = {}
        if self.oktmo_df is None or 'postal_codes' not in self.oktmo_df.columns:
            return

        for code, row in self.oktmo_df.iterrows():
            postal_codes = str(row.get('postal_codes', '')).strip()
            if postal_codes and postal_codes.lower() != 'nan':
                if '-' in postal_codes:
                    try:
                        start, end = postal_codes.split('-')
                        start = start.strip()
                        end = end.strip()
                        if start.isdigit() and end.isdigit():
                            start_int = int(start)
                            end_int = int(end)
                            for idx in range(start_int, end_int + 1):
                                self.postal_indexes[str(idx)] = str(code)
                    except ValueError:
                        if postal_codes.isdigit():
                            self.postal_indexes[postal_codes] = str(code)
                else:
                    if postal_codes.isdigit():
                        self.postal_indexes[postal_codes] = str(code)

        if self.logger:
            self.logger.info(f"Построен индекс почтовых кодов: {len(self.postal_indexes)} записей")

    def find_oktmo_by_postal(self, postal_code: str) -> Optional[str]:
        """Поиск кода ОКТМО по почтовому индексу"""
        if not postal_code or not self.postal_indexes:
            return None

        normalized_code = re.sub(r'\D', '', str(postal_code))
        if not normalized_code or len(normalized_code) != 6:
            return None

        return self.postal_indexes.get(normalized_code)

    def get_oktmo_name(self, oktmo_code: str) -> Optional[str]:
        """Получение названия по коду ОКТМО"""
        if not oktmo_code or self.oktmo_df is None:
            return None

        normalized_code = OktmoManager.normalize_oktmo_code(oktmo_code)
        if not normalized_code:
            return None

        if normalized_code in self.oktmo_df.index:
            try:
                name_value = self.oktmo_df.loc[normalized_code, 'name']
                if isinstance(name_value, pd.Series):
                    return name_value.iloc[0]
                return str(name_value).strip()
            except:
                pass

        # Поиск по префиксу (для более общих кодов)
        if len(normalized_code) < 11:
            try:
                matching_codes = [
                    code for code in self.oktmo_df.index
                    if str(code).startswith(normalized_code)
                ]
                if matching_codes:
                    name_value = self.oktmo_df.loc[matching_codes[0], 'name']
                    if isinstance(name_value, pd.Series):
                        return name_value.iloc[0]
                    return str(name_value).strip()
            except:
                pass

        return None

    def has_postal_data(self) -> bool:
        """Проверка наличия данных о почтовых индексах"""
        return bool(self.postal_indexes)

    @staticmethod
    def normalize_oktmo_code(oktmo_value: Any) -> Optional[str]:
        """Нормализация кода ОКТМО"""
        if pd.isna(oktmo_value):
            return None

        oktmo_str = re.sub(r'\D', '', str(oktmo_value).strip())
        if not oktmo_str:
            return None

        # ОКТМО может быть от 8 до 11 цифр
        if len(oktmo_str) < 8:
            oktmo_str = oktmo_str.zfill(8)
        elif len(oktmo_str) > 11:
            oktmo_str = oktmo_str[:11]

        return oktmo_str


# ============================================================================
# КЛАСС ДЛЯ РАБОТЫ СО СПРАВОЧНИКОМ НАЦИОНАЛЬНОСТЕЙ
# ============================================================================
class NationalityManager:
    """Менеджер для работы со справочником национальностей"""
    def __init__(self, logger=None):
        self.logger = logger
        self.nationality_df = None
        self.nationality_dict = {}  # Код -> Название
        self.name_to_code = {}      # Название -> Код
        
    def load_nationalities(self, file_path: Path):
        """Загрузка справочника национальностей из Excel или CSV"""
        try:
            if not file_path.exists():
                if self.logger:
                    self.logger.warning(f"Файл справочника национальностей не найден: {file_path}")
                return False

            file_extension = file_path.suffix.lower()
            
            if file_extension in ['.xlsx', '.xls', '.xlsm']:
                # Загружаем Excel
                self.nationality_df = pd.read_excel(file_path, dtype=str)
            elif file_extension == '.csv':
                # Пробуем разные кодировки для CSV
                encodings = ['utf-8-sig', 'cp1251', 'windows-1251', 'utf-8']
                for encoding in encodings:
                    try:
                        self.nationality_df = pd.read_csv(
                            file_path, 
                            dtype=str,
                            encoding=encoding,
                            on_bad_lines='skip'
                        )
                        break
                    except:
                        if encoding == encodings[-1]:
                            raise
                        continue
            else:
                if self.logger:
                    self.logger.error(f"Неподдерживаемый формат файла национальностей: {file_extension}")
                return False

            # Поиск колонок с кодами и названиями
            code_col = None
            name_col = None
            
            for col in self.nationality_df.columns:
                col_lower = str(col).lower()
                if any(word in col_lower for word in ['код', 'code', 'num', 'номер']):
                    code_col = col
                elif any(word in col_lower for word in ['наимен', 'name', 'название', 'nationality']):
                    name_col = col
            
            # Если не нашли стандартные названия, берем первые две колонки
            if code_col is None and len(self.nationality_df.columns) > 0:
                code_col = self.nationality_df.columns[0]
            if name_col is None and len(self.nationality_df.columns) > 1:
                name_col = self.nationality_df.columns[1]
            
            if code_col is None:
                if self.logger:
                    self.logger.error("Не найдена колонка с кодами национальностей")
                return False
            
            # Создаем словари
            for idx, row in self.nationality_df.iterrows():
                code = str(row[code_col]).strip() if code_col in row else ""
                name = str(row[name_col]).strip() if name_col in row and name_col in row else ""
                
                if code:
                    self.nationality_dict[code] = name
                    if name:
                        self.name_to_code[name.lower()] = code
            
            if self.logger:
                self.logger.info(f"Загружено национальностей: {len(self.nationality_dict)}")
                if len(self.nationality_dict) > 0:
                    sample = list(self.nationality_dict.items())[:3]
                    self.logger.info(f"Пример национальностей: {sample}")
            
            return True
            
        except Exception as e:
            if self.logger:
                self.logger.error(f"Ошибка загрузки справочника национальностей: {e}", exc_info=True)
            return False
    
    def get_nationality_by_code(self, code: str) -> Optional[str]:
        """Получение названия национальности по коду"""
        if not code or not self.nationality_dict:
            return None
        
        normalized_code = str(code).strip()
        return self.nationality_dict.get(normalized_code)
    
    def get_code_by_nationality(self, name: str) -> Optional[str]:
        """Получение кода национальности по названию"""
        if not name or not self.name_to_code:
            return None
        
        normalized_name = str(name).strip().lower()
        return self.name_to_code.get(normalized_name)
    
    def has_data(self) -> bool:
        """Проверка наличия данных"""
        return bool(self.nationality_dict)


# ============================================================================
# АДРЕСНЫЙ ОБОГАТИТЕЛЬ
# ============================================================================
class AddressEnricher:
    """Класс для обогащения адресных данных"""
    def __init__(self, oktmo_manager: OktmoManager, nationality_manager: NationalityManager = None, logger=None):
        self.oktmo_manager = oktmo_manager
        self.nationality_manager = nationality_manager
        self.logger = logger
        self.cache = {}

    def print_stats(self):
        """Вывод статистики по справочникам"""
        if self.logger:
            self.logger.info(f"Справочник ОКТМО: {len(self.oktmo_manager.oktmo_df)} записей")
            self.logger.info(f"Почтовые индексы: {len(self.oktmo_manager.postal_indexes)} записей")
            if self.oktmo_manager.has_postal_data():
                self.logger.info("Поиск почтовых индексов: ДОСТУПЕН")
            else:
                self.logger.info("Поиск почтовых индексов: НЕДОСТУПЕН (загрузите справочник)")
            
            if self.nationality_manager and self.nationality_manager.has_data():
                self.logger.info(f"Справочник национальностей: {len(self.nationality_manager.nationality_dict)} записей")
            else:
                self.logger.info("Справочник национальностей: НЕ ЗАГРУЖЕН")

    def enrich_postal_code(self, address: str, current_code: Optional[str] = None) -> Optional[str]:
        """Обогащение почтового индекса на основе ОКТМО"""
        if not self.oktmo_manager.has_postal_data():
            if self.logger:
                self.logger.warning("Попытка поиска индекса без справочника ОКТМО")
            return current_code

        address_str = str(address).strip()
        postal_match = re.search(r'\b\d{6}\b', address_str)
        if postal_match:
            found_code = postal_match.group(0)
            if self.logger:
                self.logger.info(f"Найден индекс в адресе: {found_code}")
            return found_code

        return current_code

    def enrich_oktmo_from_postal(self, postal_code: str) -> Optional[str]:
        """Поиск кода ОКТМО по почтовому индексу"""
        if not postal_code or not self.oktmo_manager.has_postal_data():
            return None

        oktmo_code = self.oktmo_manager.find_oktmo_by_postal(postal_code)
        if oktmo_code and self.logger:
            self.logger.info(f"Найден ОКТМО {oktmo_code} по индексу {postal_code}")
            return oktmo_code

        return None

    def enrich_municipality_code(self, raw_data: Dict[str, Any], df_row: pd.Series = None) -> Tuple[Optional[str], Optional[str]]:
        """Обогащение кода муниципального образования"""
        if not self.oktmo_manager.oktmo_df.empty:
            # Проверяем наличие кода в данных
            current_code = raw_data.get('municipality_code')
            if pd.isna(current_code) or not str(current_code).strip():
                # Если кода нет, пытаемся найти по почтовому индексу
                postal_code = raw_data.get('postal_code')
                if postal_code and not pd.isna(postal_code):
                    oktmo_code = self.enrich_oktmo_from_postal(str(postal_code))
                    if oktmo_code:
                        oktmo_name = self.oktmo_manager.get_oktmo_name(oktmo_code)
                        # Обогащаем исходный DataFrame если передан
                        if df_row is not None and 'municipality_code' in df_row.index:
                            df_row['municipality_code'] = oktmo_code
                        return oktmo_code, oktmo_name

            # Если код есть, получаем его название
            elif current_code:
                oktmo_name = self.oktmo_manager.get_oktmo_name(str(current_code))
                return str(current_code).strip(), oktmo_name

        return None, None

    def enrich_nationality(self, raw_data: Dict[str, Any], df_row: pd.Series = None) -> Tuple[Optional[str], Optional[str]]:
        """Обогащение национальности"""
        if not self.nationality_manager or not self.nationality_manager.has_data():
            return None, None
        
        nationality_name = raw_data.get('nationality_name', '')
        nationality_code = raw_data.get('nationality_code', '')
        
        # Если есть код, получаем название
        if nationality_code and str(nationality_code).strip():
            name = self.nationality_manager.get_nationality_by_code(str(nationality_code).strip())
            if name:
                return str(nationality_code).strip(), name
        
        # Если есть название, получаем код
        elif nationality_name and str(nationality_name).strip():
            code = self.nationality_manager.get_code_by_nationality(str(nationality_name).strip())
            if code:
                return code, str(nationality_name).strip()
        
        return None, None


# ============================================================================
# ВАЛИДАТОРЫ И ПРОЦЕССОРЫ
# ============================================================================
class FieldValidator:
    """Валидация полей"""
    
    @staticmethod
    def validate_date_format(date_str: str, field_name: str = "Дата") -> Tuple[bool, Optional[str]]:
        """Строгая проверка формата даты ГГГГ-ММ-ДД"""
        if not date_str or str(date_str).strip() == "":
            return True, None  # Пустые значения допустимы
        
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        if not re.match(pattern, str(date_str).strip()):
            return False, f"{field_name}: неверный формат (требуется ГГГГ-ММ-ДД)"
        
        try:
            year, month, day = map(int, str(date_str).strip().split('-'))
            if month < 1 or month > 12:
                return False, f"{field_name}: неверный месяц (1-12)"
            if day < 1 or day > 31:
                return False, f"{field_name}: неверный день (1-31)"
            
            # Проверка на корректную дату
            datetime(year, month, day)
            return True, None
        except ValueError as e:
            return False, f"{field_name}: некорректная дата ({str(e)})"
    
    @staticmethod
    def validate_normalized_date(date_str: str, field_name: str = "Дата",
                                 check_future: bool = True,
                                 min_year: int = 1900,
                                 max_year: int = 2050) -> Tuple[bool, Optional[str]]:
        """Валидация даты с проверкой диапазона"""
        if not date_str or str(date_str).strip() == "":
            return True, None
        
        # Сначала проверяем формат
        is_valid, error = FieldValidator.validate_date_format(date_str, field_name)
        if not is_valid:
            return False, error
        
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            if dt.year < min_year or dt.year > max_year:
                return False, f"{field_name}: год вне диапазона ({min_year}-{max_year})"
            if check_future and dt > datetime.now():
                return False, f"{field_name}: не может быть в будущем"
            return True, None
        except Exception:
            return False, f"{field_name}: ошибка валидации даты"

    @staticmethod
    def validate_text(text: str, field_name: str = "Поле",
                      min_length: int = 1,
                      max_length: int = 500,
                      allow_empty: bool = True) -> Tuple[bool, Optional[str]]:
        """Валидация текстового поля"""
        if not text and allow_empty:
            return True, None
        
        text_str = str(text).strip()
        if len(text_str) < min_length:
            return False, f"{field_name}: минимум {min_length} символов"
        if len(text_str) > max_length:
            return False, f"{field_name}: максимум {max_length} символов"
        return True, None

    @staticmethod
    def validate_numeric_code(code: str, field_name: str = "Код",
                              exact_length: Optional[int] = None,
                              min_length: int = 1,
                              max_length: int = 20,
                              allow_empty: bool = True) -> Tuple[bool, Optional[str]]:
        """Валидация числового кода"""
        if not code and allow_empty:
            return True, None
        
        code_str = str(code).strip()
        if not code_str.isdigit():
            return False, f"{field_name}: только цифры"
        if exact_length and len(code_str) != exact_length:
            return False, f"{field_name}: должно быть {exact_length} цифр"
        if len(code_str) < min_length or len(code_str) > max_length:
            return False, f"{field_name}: от {min_length} до {max_length} цифр"
        return True, None

    @staticmethod
    def validate_region_code(code: str) -> Tuple[bool, Optional[str]]:
        """Валидация кода субъекта РФ (2 цифры)"""
        if not code or str(code).strip() == "":
            return True, None
        
        code_str = str(code).strip()
        if not code_str.isdigit() or len(code_str) != 2:
            return False, "Код региона: должен быть 2 цифры"
        if code_str not in RUSSIAN_REGIONS:
            return False, f"Код региона: неизвестный код {code_str}"
        return True, None

    @staticmethod
    def validate_snils(snils: str) -> Tuple[bool, Optional[str]]:
        """Валидация СНИЛС (XXX-XXX-XXX XX)"""
        if not snils or str(snils).strip() == "":
            return True, None
        
        snils_str = str(snils).strip()
        pattern = r'^\d{3}-\d{3}-\d{3}\s\d{2}$'
        if not re.match(pattern, snils_str):
            return False, "СНИЛС: неверный формат (должен быть XXX-XXX-XXX XX)"
        
        # Проверка контрольного числа (упрощенная)
        try:
            digits = re.sub(r'\D', '', snils_str)
            if len(digits) != 11:
                return False, "СНИЛС: должно быть 11 цифр"
        except:
            return False, "СНИЛС: ошибка проверки"
        
        return True, None

    @staticmethod
    def validate_ticket_series(series: str) -> Tuple[bool, Optional[str]]:
        """Валидация серии охотничьего билета (до 4 символов, цифры/буквы)"""
        if not series or str(series).strip() == "":
            return True, None
        
        series_str = str(series).strip()
        if len(series_str) > 4:
            return False, "Серия билета: максимум 4 символа"
        
        # Разрешаем буквы и цифры
        if not re.match(r'^[A-Za-zА-Яа-я0-9]*$', series_str):
            return False, "Серия билета: только буквы и цифры"
        
        return True, None

    @staticmethod
    def validate_ticket_number(number: str) -> Tuple[bool, Optional[str]]:
        """Валидация номера охотничьего билета (до 6 цифр)"""
        if not number or str(number).strip() == "":
            return True, None
        
        number_str = str(number).strip()
        if len(number_str) > 6:
            return False, "Номер билета: максимум 6 цифр"
        if not number_str.isdigit():
            return False, "Номер билета: только цифры"
        return True, None

    @staticmethod
    def validate_passport_series(series: str) -> Tuple[bool, Optional[str]]:
        """Валидация серии паспорта (до 4 цифр)"""
        if not series or str(series).strip() == "":
            return True, None
        
        series_str = str(series).strip()
        if len(series_str) > 4:
            return False, "Серия паспорта: максимум 4 цифры"
        if not series_str.isdigit():
            return False, "Серия паспорта: только цифры"
        return True, None

    @staticmethod
    def validate_passport_number(number: str) -> Tuple[bool, Optional[str]]:
        """Валидация номера паспорта (до 6 цифр)"""
        if not number or str(number).strip() == "":
            return True, None
        
        number_str = str(number).strip()
        if len(number_str) > 6:
            return False, "Номер паспорта: максимум 6 цифр"
        if not number_str.isdigit():
            return False, "Номер паспорта: только цифры"
        return True, None

    @staticmethod
    def validate_phone(phone: str) -> Tuple[bool, Optional[str]]:
        """Валидация телефона в формате +7XXXXXXXXXX"""
        if not phone or str(phone).strip() == "":
            return True, None
        
        phone_str = str(phone).strip()
        pattern = r'^\+7\d{10}$'
        if not re.match(pattern, phone_str):
            return False, "Телефон: неверный формат (требуется +7XXXXXXXXXX)"
        
        return True, None

    @staticmethod
    def validate_indigenous_people(value: str) -> Tuple[bool, Optional[str]]:
        """Валидация поля коренных народов"""
        if not value or str(value).strip() == "":
            return True, None
        
        val_str = str(value).strip().lower()
        if val_str not in ["true", "false", ""]:
            return False, "Коренные народы: должно быть 'true', 'false' или пусто"
        return True, None


# Обновленные правила валидации
FIELD_VALIDATION_RULES = {
    'date_entry': {
        'validator': FieldValidator.validate_date_format,
        'params': {'field_name': 'Дата внесения'}
    },
    'date_issue': {
        'validator': FieldValidator.validate_date_format,
        'params': {'field_name': 'Дата выдачи'}
    },
    'date_issue_ticket': {
        'validator': FieldValidator.validate_date_format,
        'params': {'field_name': 'Дата выдачи билета'}
    },
    'birth_date': {
        'validator': FieldValidator.validate_normalized_date,
        'params': {'field_name': 'Дата рождения', 'check_future': True, 'min_year': 1900}
    },
    'municipality_code': {
        'validator': FieldValidator.validate_numeric_code,
        'params': {'field_name': 'Код МО', 'min_length': 6, 'max_length': 11}
    },
    'surname': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Фамилия', 'min_length': 2, 'max_length': 100}
    },
    'hunter_name': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Имя', 'min_length': 2, 'max_length': 100}
    },
    'patronymic': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Отчество', 'min_length': 2, 'max_length': 100, 'allow_empty': True}
    },
    'birth_place': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Место рождения', 'min_length': 2, 'max_length': 500}
    },
    'postal_address': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Почтовый адрес', 'min_length': 5, 'max_length': 500, 'allow_empty': True}
    },
    'postal_code': {
        'validator': FieldValidator.validate_numeric_code,
        'params': {'field_name': 'Индекс', 'exact_length': 6, 'allow_empty': True}
    },
    'address': {
        'validator': FieldValidator.validate_text,
        'params': {'field_name': 'Адрес', 'min_length': 5, 'max_length': 500, 'allow_empty': True}
    },
    'snils_code': {
        'validator': FieldValidator.validate_snils,
        'params': {}
    },
    'series_passport': {
        'validator': FieldValidator.validate_passport_series,
        'params': {}
    },
    'number_passport': {
        'validator': FieldValidator.validate_passport_number,
        'params': {}
    },
    'series_ticket': {
        'validator': FieldValidator.validate_ticket_series,
        'params': {}
    },
    'number_ticket': {
        'validator': FieldValidator.validate_ticket_number,
        'params': {}
    },
    'region_code': {
        'validator': FieldValidator.validate_region_code,
        'params': {}
    },
    'phone': {
        'validator': FieldValidator.validate_phone,
        'params': {}
    },
    'is_belonged_to_indigenous_people': {
        'validator': FieldValidator.validate_indigenous_people,
        'params': {}
    },
}


class UniversalDateParser:
    @staticmethod
    def parse(value: Any) -> Tuple[Optional[datetime], str]:
        if pd.isna(value) or str(value).strip() == '':
            return None, ''

        val_str = re.sub(r'\s+', ' ', str(value).strip())

        if isinstance(value, (int, float)) and 10000 < value < 100000:
            try:
                dt = pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(value))
                return dt.to_pydatetime(), dt.strftime('%Y-%m-%d')
            except:
                pass

        if 'e+' in val_str.lower():
            return None, val_str

        if isinstance(value, (pd.Timestamp, datetime)):
            dt = value if isinstance(value, datetime) else value.to_pydatetime()
            return dt, dt.strftime('%Y-%m-%d')

        try:
            dtobj = pd.to_datetime(val_str, errors='coerce')
            if pd.notna(dtobj):
                return dtobj.to_pydatetime(), dtobj.strftime('%Y-%m-%d')
            else:
                # Попробуем парсить как timestamp (секунды)
                try:
                    ts = float(val_str)
                    if ts > 1000000000 and ts < 2000000000:  # Примерно 2001-2033 год
                        dt = datetime.fromtimestamp(ts)
                        return dt, dt.strftime('%Y-%m-%d')
                except:
                    pass
        except:
            pass

        date_formats = ['%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y',
                        '%Y.%m.%d', '%d.%m.%y', '%d-%m-%y', '%d/%m/%y',
                        '%Y-%m-%d %H:%M:%S', '%d.%m.%Y %H:%M:%S']

        for fmt in date_formats:
            try:
                dt = datetime.strptime(val_str, fmt)
                return dt, dt.strftime('%Y-%m-%d')
            except:
                continue

        return None, val_str


class ChangeLogger:
    def __init__(self):
        self.changes = defaultdict(list)
        self.errors = defaultdict(list)
        self.enrichments = defaultdict(list)
        self.missing_required = defaultdict(list)
        self.skipped_rows = []

    def log_change(self, row: int, field: str, original: Any, converted: Any):
        self.changes[row].append({
            'field': field,
            'original': str(original)[:100] if pd.notna(original) else '',
            'converted': str(converted)[:100] if converted is not None else ''
        })

    def log_error(self, row: int, field: str, value: Any, error: str):
        self.errors[row].append({
            'field': field,
            'value': str(value)[:100] if pd.notna(value) else '',
            'error': error
        })

    def log_enrichment(self, row: int, field: str, old_value: Any, new_value: Any):
        self.enrichments[row].append({
            'field': field,
            'original': str(old_value)[:100] if old_value else '',
            'converted': str(new_value)[:100] if new_value else ''
        })

    def log_missing_required(self, row: int, field: str, value: Any = None):
        self.missing_required[row].append({
            'field': field,
            'value': str(value)[:100] if pd.notna(value) else ''
        })

    def log_skipped_row(self, row: int, reason: str):
        self.skipped_rows.append({
            'row': row,
            'reason': reason
        })

    def get_summary(self) -> Dict:
        return {
            'total_changes': len(self.changes),
            'total_errors': len(self.errors),
            'total_enrichments': len(self.enrichments),
            'total_missing_required': sum(len(fields) for fields in self.missing_required.values()),
            'total_skipped_rows': len(self.skipped_rows),
            'changes': dict(self.changes),
            'errors': dict(self.errors),
            'enrichments': dict(self.enrichments),
            'missing_required': dict(self.missing_required),
            'skipped_rows': self.skipped_rows
        }


class DataProcessor:
    def __init__(self, mode: str = 'smart', enable_logging: bool = False,
                 address_enricher: Optional[AddressEnricher] = None,
                 enrich_postal: bool = False, enrich_oktmo: bool = False,
                 selected_region: str = None, logger=None):
        self.mode = mode
        self.date_parser = UniversalDateParser()
        self.logger = logger
        self.logger_obj = ChangeLogger() if enable_logging else None
        self.address_enricher = address_enricher
        self.enrich_postal = enrich_postal
        self.enrich_oktmo = enrich_oktmo
        self.selected_region = selected_region

        if self.enrich_postal and self.address_enricher and not self.address_enricher.oktmo_manager.has_postal_data():
            if self.logger:
                self.logger.warning("Поиск почтовых индексов невозможен: не загружен справочник ОКТМО")

    def normalize_snils(self, value: Any) -> str:
        """Нормализация СНИЛС"""
        if pd.isna(value):
            return ""
        snils_str = str(value).strip()
        
        # Удаляем все нецифровые символы
        digits = re.sub(r'\D', '', snils_str)
        
        if len(digits) == 11:
            return f"{digits[0:3]}-{digits[3:6]}-{digits[6:9]} {digits[9:11]}"
        
        # Если уже в правильном формате, возвращаем как есть
        if re.match(r'^\d{3}-\d{3}-\d{3}\s\d{2}$', snils_str):
            return snils_str
        
        return snils_str if snils_str else ""

    def normalize_phone(self, value: Any) -> str:
        """Нормализация телефона - преобразование в формат +7XXXXXXXXXX"""
        if pd.isna(value):
            return ""

        phone_str = str(value).strip()

        # Обработка научной нотации
        if 'e+' in phone_str.lower():
            try:
                phone_num = int(float(phone_str))
                phone_str = str(phone_num)
            except:
                return phone_str if phone_str else ""

        # Удаляем все нецифровые символы
        digits = re.sub(r'\D', '', phone_str)

        if not digits:
            return phone_str if phone_str else ""

        # Приводим к формату +7XXXXXXXXXX
        if digits.startswith('7') or digits.startswith('8'):
            # Убираем первую цифру (7 или 8) и добавляем +7
            if len(digits) >= 11:
                return f"+7{digits[-10:]}"
            else:
                # Если цифр меньше 11, дополняем до 10 цифр
                return f"+7{digits[-9:].zfill(10)}"
        elif digits.startswith('9') and len(digits) == 10:
            # Если номер начинается с 9 и 10 цифр - это мобильный РФ
            return f"+7{digits}"
        elif len(digits) == 10:
            # Если просто 10 цифр - добавляем +7
            return f"+7{digits}"
        elif len(digits) == 11 and digits.startswith('7'):
            # Если 11 цифр и начинается с 7
            return f"+7{digits[1:11]}"
        else:
            # В остальных случаях возвращаем все цифры с +7
            if digits:
                return f"+7{digits[-10:]}"
            return digits

    def normalize_postal_code(self, value: Any) -> str:
        """Нормализация почтового индекса"""
        if not value or pd.isna(value):
            return ""
        code_str = str(value).strip()
        digits = re.sub(r'\D', '', code_str)
        if len(digits) == 6:
            return digits
        return code_str if code_str else ""

    def normalize_indigenous_people(self, value: Any) -> str:
        """Нормализация поля коренных народов"""
        if pd.isna(value):
            return ""
        
        val_str = str(value).strip().lower()
        
        if val_str in ["true", "1", "да", "yes", "истина"]:
            return "true"
        elif val_str in ["false", "0", "нет", "no", "ложь"]:
            return "false"
        elif val_str == "":
            return ""
        else:
            # Для нераспознанных значений возвращаем "false"
            return "false"

    def convert_to_string(self, value: Any) -> str:
        """Конвертация значения в строку, удаление дробной части если это целое число"""
        if pd.isna(value):
            return ""

        # Если значение float и оно целое, преобразуем в int
        if isinstance(value, float):
            if value.is_integer():
                value = int(value)
        
        str_value = str(value).strip()
        if not str_value:
            return ""

        # Убираем пробелы и возвращаем как строку
        return str_value.replace(" ", "")

    def process_row(self, row_idx: int, raw: Dict[str, Any], df_row: pd.Series = None) -> Tuple[Optional[Dict[str, Any]], Optional[Dict]]:
        """Обработка строки данных с пропуском строк без обязательных полей"""
        result = {}

        # Обработка полей
        for field, value in raw.items():
            try:
                # Всегда добавляем поле, даже если оно пустое
                if value is None or pd.isna(value) or (isinstance(value, str) and str(value).strip() == ''):
                    result[field] = ""
                    continue

                # Если значение - Series, берем первое значение
                if isinstance(value, pd.Series):
                    if value.empty:
                        result[field] = ""
                        continue
                    value = value.iloc[0] if len(value) > 0 else value

                # Обработка дат
                if 'date' in field.lower() or 'issue' in field.lower():
                    dt_obj, normalized = self.date_parser.parse(value)
                    if dt_obj:
                        result[field] = normalized
                        if str(value).strip() != normalized:
                            if self.logger_obj:
                                self.logger_obj.log_change(row_idx, field, value, normalized)
                    else:
                        result[field] = str(value).strip() if str(value).strip() else ""

                # Обработка СНИЛС
                elif 'snils' in field.lower():
                    normalized_snils = self.normalize_snils(value)
                    result[field] = normalized_snils
                    if str(value).strip() != normalized_snils:
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, normalized_snils)

                # Обработка телефонов
                elif 'phone' in field.lower() or 'tel' in field.lower():
                    normalized_phone = self.normalize_phone(value)
                    result[field] = normalized_phone
                    if str(value).strip() != normalized_phone:
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, normalized_phone)

                # Обработка почтовых индексов
                elif 'postal' in field.lower() and 'code' in field.lower():
                    normalized_postal = self.normalize_postal_code(value)
                    result[field] = normalized_postal
                    if str(value).strip() != normalized_postal:
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, normalized_postal)

                # Обработка коренных народов
                elif 'indigenous' in field.lower():
                    normalized_indigenous = self.normalize_indigenous_people(value)
                    result[field] = normalized_indigenous
                    if str(value).strip() != normalized_indigenous:
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, normalized_indigenous)

                # Обработка числовых полей - конвертируем в строку с удалением дробной части
                elif any(x in field.lower() for x in ['series', 'number', 'code']):
                    # Специальная обработка для целых чисел
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    string_value = self.convert_to_string(value)
                    result[field] = string_value
                    if str(value).strip() != str(string_value):
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, string_value)

                # Общая обработка текста
                else:
                    # Обработка float значений для других полей
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    cleaned = re.sub(r'\s+', ' ', str(value)).strip()
                    result[field] = cleaned if cleaned else ""
                    if str(value).strip() != cleaned:
                        if self.logger_obj:
                            self.logger_obj.log_change(row_idx, field, value, cleaned)

            except Exception as e:
                result[field] = str(value).strip() if str(value).strip() else ""
                if self.logger:
                    self.logger.warning(f"Ошибка обработки поля {field} в строке {row_idx}: {e}")

        
        # Обогащение муниципального кода из ОКТМО
        if self.address_enricher and self.enrich_oktmo:
            municipality_code, municipality_name = self.address_enricher.enrich_municipality_code(result, df_row)
            if municipality_code:
                result['municipality_code'] = municipality_code
                if municipality_name:
                    result['municipality_name'] = municipality_name
                    if self.logger_obj:
                        self.logger_obj.log_enrichment(
                            row_idx,
                            'municipality',
                            None,
                            f"{municipality_code} - {municipality_name}"
                        )

        # Обогащение национальности
        if self.address_enricher and self.address_enricher.nationality_manager:
            nationality_code, nationality_name = self.address_enricher.enrich_nationality(result, df_row)
            if nationality_code or nationality_name:
                if nationality_code:
                    result['nationality_code'] = nationality_code
                if nationality_name:
                    result['nationality_name'] = nationality_name
                    if self.logger_obj:
                        self.logger_obj.log_enrichment(
                            row_idx,
                            'nationality',
                            result.get('nationality_name', ''),
                            nationality_name
                        )

        # Валидация полей по правилам (кроме region_code)
        for field, rules in FIELD_VALIDATION_RULES.items():
            if field in result and field != 'region_code':
                value = result[field]
                if pd.isna(value) or value == "":
                    continue

                validator = rules['validator']
                params = rules['params']

                is_valid, error_msg = validator(str(value), **params)
                if not is_valid:
                    if self.logger_obj:
                        self.logger_obj.log_error(row_idx, field, value, error_msg)
                    if self.logger:
                        self.logger.warning(f"Строка {row_idx}: {error_msg}")

        return result, self.logger_obj.get_summary() if self.logger_obj else None

    def filter_by_region(self, row: Dict[str, Any]) -> bool:
        """Фильтрация строк по региону"""
        if not self.selected_region:
            return True

        region_code = row.get('region_code')
        if pd.isna(region_code) or not region_code:
            return True

        return str(region_code).strip() == str(self.selected_region).zfill(2)


# ============================================================================
# ОСНОВНОЙ КЛАСС КОНВЕРТЕРА
# ============================================================================
class ExcelToJsonConverter:
    def __init__(self):
        self.logger = None
        self.log_file = None
        self.oktmo_manager = None
        self.nationality_manager = None
        self.address_enricher = None
        self.progress_callback = None
        self.start_time = None
        self.processed_rows = 0
        self.total_rows = 0
        self.enriched_data = None

    def set_progress_callback(self, callback):
        """Установка коллбэка для обновления прогресса"""
        self.progress_callback = callback

    def _update_progress(self, current, total, message="", time_remaining="--:--"):
        """Обновление прогресса с расчетом времени"""
        if self.progress_callback:
            if total > 0:
                percent = (current / total) * 100
            else:
                percent = 0

            if time_remaining == "--:--" and self.start_time and current > 0 and current < total:
                elapsed = time.time() - self.start_time
                rows_per_second = current / elapsed
                if rows_per_second > 0:
                    remaining_rows = total - current
                    remaining_time = remaining_rows / rows_per_second
                    minutes = int(remaining_time // 60)
                    seconds = int(remaining_time % 60)
                    time_remaining = f"{minutes:02d}:{seconds:02d}"

            self.progress_callback(percent, message, time_remaining)

    def _log_gui_and_console(self, message: str, level: str, gui_callback):
        """Логирование в GUI и консоль"""
        if gui_callback:
            gui_callback.put((message, level))
        if self.logger:
            if level == "ERROR":
                self.logger.error(message)
            elif level == "WARNING":
                self.logger.warning(message)
            elif level == "SUCCESS":
                self.logger.info(f"✓ {message}")
            else:
                self.logger.info(message)

    def convert(self,
                input_file: Path,
                output_folder: Path,
                sheet_name: str = None,
                mode: str = 'smart',
                create_report: bool = False,
                split_count: int = None,
                include_postal: bool = False,
                include_oktmo: bool = False,
                region_code: int = None,
                oktmo_csv_path: Optional[Path] = None,
                nationality_file: Optional[Path] = None,
                gui_callback = None,
                selected_region: str = None) -> Dict[str, Any]:
        """
        Основной метод конвертации с созданием двух JSON файлов
        """
        # Настройка логирования
        self.logger, self.log_file = setup_logging(input_file)

        self._log_gui_and_console("🚀 Начало конвертации", "INFO", gui_callback)
        self._log_gui_and_console(f"📁 Входной файл: {input_file}", "INFO", gui_callback)
        self._log_gui_and_console(f"📁 Выходная папка: {output_folder}", "INFO", gui_callback)
        self._log_gui_and_console(f"⚙️ Режим: {mode}", "INFO", gui_callback)

        if selected_region:
            self._log_gui_and_console(f"📍 Выбранный регион: {selected_region}", "INFO", gui_callback)
        else:
            self._log_gui_and_console("📍 Регион не выбран", "INFO", gui_callback)

        if not input_file.exists():
            error_msg = f"❌ Входной файл не найден: {input_file}"
            self._log_gui_and_console(error_msg, "ERROR", gui_callback)
            raise FileNotFoundError(error_msg)

        try:
            self.start_time = time.time()

            # Загрузка данных из Excel/CSV
            self._log_gui_and_console("📥 Загрузка данных...", "INFO", gui_callback)
            self._update_progress(0, 100, "Загрузка данных...")
            df = self._load_data(input_file, sheet_name)

            self._log_gui_and_console(f"📊 Загружено строк: {len(df)}", "INFO", gui_callback)
            self._log_gui_and_console(f"📊 Загружено столбцов: {len(df.columns)}", "INFO", gui_callback)

            # Проверяем обязательные поля
            missing_required = []
            for field in REQUIRED_FIELDS:
                if field not in df.columns:
                    missing_required.append(field)

            if missing_required:
                self._log_gui_and_console(f"⚠️ ВНИМАНИЕ: Отсутствуют обязательные поля: {', '.join(missing_required)}", "WARNING", gui_callback)

            if df.empty:
                error_msg = "❌ Файл пуст или не содержит данных"
                self._log_gui_and_console(error_msg, "ERROR", gui_callback)
                raise ValueError(error_msg)

            # Сохраняем копию для обогащения
            self.enriched_data = df.copy()

            # Инициализация менеджеров
            self.oktmo_manager = OktmoManager(self.logger)
            self.nationality_manager = NationalityManager(self.logger)
            self.address_enricher = AddressEnricher(self.oktmo_manager, self.nationality_manager, self.logger)

            # Загрузка внешнего справочника ОКТМО
            oktmo_loaded = False
            if oktmo_csv_path and oktmo_csv_path.exists():
                self._log_gui_and_console(f"📥 Загрузка справочника ОКТМО из: {oktmo_csv_path}", "INFO", gui_callback)
                self._update_progress(5, 100, "Загрузка справочника ОКТМО...")
                if self.oktmo_manager.load_external_oktmo(oktmo_csv_path):
                    self._log_gui_and_console("✅ Справочник ОКТМО успешно загружен", "SUCCESS", gui_callback)
                    oktmo_loaded = True
                else:
                    self._log_gui_and_console("⚠️ Не удалось загрузить справочник ОКТМО", "WARNING", gui_callback)
            else:
                oktmo_file = input_file.parent / "oktmo.csv"
                if oktmo_file.exists():
                    self._log_gui_and_console(f"🔍 Автопоиск справочника ОКТМО: {oktmo_file}", "INFO", gui_callback)
                    self._update_progress(5, 100, "Загрузка справочника ОКТМО...")
                    if self.oktmo_manager.load_external_oktmo(oktmo_file):
                        self._log_gui_and_console("✅ Справочник ОКТМО успешно загружен", "SUCCESS", gui_callback)
                        oktmo_loaded = True
                    else:
                        self._log_gui_and_console("⚠️ Не удалось загрузить справочник ОКТМО", "WARNING", gui_callback)

            # Загрузка справочника национальностей
            nationality_loaded = False
            if nationality_file and nationality_file.exists():
                self._log_gui_and_console(f"📥 Загрузка справочника национальностей из: {nationality_file}", "INFO", gui_callback)
                self._update_progress(7, 100, "Загрузка справочника национальностей...")
                if self.nationality_manager.load_nationalities(nationality_file):
                    self._log_gui_and_console("✅ Справочник национальностей успешно загружен", "SUCCESS", gui_callback)
                    nationality_loaded = True
                else:
                    self._log_gui_and_console("⚠️ Не удалось загрузить справочник национальностей", "WARNING", gui_callback)

            if oktmo_loaded or nationality_loaded:
                self.address_enricher.print_stats()

            if include_postal and not self.oktmo_manager.has_postal_data():
                warning_msg = "⚠️ Поиск почтовых индексов невозможен: не загружен справочник ОКТМО"
                self._log_gui_and_console(warning_msg, "WARNING", gui_callback)
                include_postal = False

            # Создание процессора данных
            data_processor = DataProcessor(
                mode=mode,
                enable_logging=True,
                address_enricher=self.address_enricher,
                enrich_postal=include_postal,
                enrich_oktmo=include_oktmo,
                selected_region=str(region_code) if region_code else None,
                logger=self.logger
            )

            # Обработка данных
            self._log_gui_and_console("🔄 Обработка данных...", "INFO", gui_callback)
            self._update_progress(10, 100, "Обработка данных...")
            hunters_data = []
            tickets_data = []
            self.total_rows = len(df)
            skipped_rows = 0
            self.processed_rows = 0
            empty_rows_count = 0

            for idx, row in df.iterrows():
                row_idx = idx + 1

                # Проверка на две пустые строки подряд
                all_empty = True
                for field in df.columns:
                    if field in row:
                        value = row[field]
                        if isinstance(value, pd.Series):
                            if not value.empty and not value.isna().all() and not (value.astype(str).str.strip() == '').all():
                                all_empty = False
                                break
                        elif not pd.isna(value) and str(value).strip() != '':
                            all_empty = False
                            break

                if all_empty:
                    empty_rows_count += 1
                    if empty_rows_count >= 2:
                        self._log_gui_and_console(f"⚠️ Обнаружены 2 пустые строки подряд (строка {row_idx}). Обработка остановлена.", "WARNING", gui_callback)
                        break
                    continue
                else:
                    empty_rows_count = 0

                # Обновление прогресса
                if idx % 100 == 0 or idx == self.total_rows - 1:
                    self._update_progress(idx, self.total_rows, f"Обработка строки {idx+1} из {self.total_rows}")

                # Обработка строки
                processed_row, _ = data_processor.process_row(row_idx, row.to_dict(), 
                                                             self.enriched_data.iloc[idx] if self.enriched_data is not None else None)

                if processed_row is None:
                    skipped_rows += 1
                    continue

                # Преобразование в требуемый формат JSON
                formatted_hunter = self._format_hunter_to_json(processed_row, selected_region)
                hunters_data.append(formatted_hunter)

                # Создание записи для huntingtickets.json
                formatted_ticket = self._format_ticket_to_json(processed_row, row_idx)
                tickets_data.append(formatted_ticket)

                self.processed_rows += 1

            self._log_gui_and_console(f"✅ Обработано строк: {self.processed_rows}", "INFO", gui_callback)
            self._log_gui_and_console(f"⏭️ Пропущено строк: {skipped_rows}", "INFO", gui_callback)

            if not hunters_data:
                self._log_gui_and_console("⚠️ ВНИМАНИЕ: Нет данных для конвертации!", "WARNING", gui_callback)
                self._log_gui_and_console("🔍 Проверьте наличие обязательных полей в файле", "INFO", gui_callback)

            # Создание выходной директории
            output_folder.mkdir(parents=True, exist_ok=True)

            # Сохранение hunters.json
            self._log_gui_and_console("💾 Сохранение hunters.json...", "INFO", gui_callback)
            self._update_progress(90, 100, "Сохранение hunters.json...")
            hunters_file = output_folder / "hunters.json"
            self._save_json(hunters_data, hunters_file)
            self._log_gui_and_console(f"✅ Файл сохранен: {hunters_file}", "INFO", gui_callback)

            # Сохранение huntingtickets.json
            self._log_gui_and_console("💾 Сохранение huntingtickets.json...", "INFO", gui_callback)
            self._update_progress(95, 100, "Сохранение huntingtickets.json...")
            tickets_file = output_folder / "huntingtickets.json"
            self._save_json(tickets_data, tickets_file)
            self._log_gui_and_console(f"✅ Файл сохранен: {tickets_file}", "INFO", gui_callback)

            # Создание отчета
            report_file = None
            if create_report:
                self._log_gui_and_console("📊 Создание отчета...", "INFO", gui_callback)
                self._update_progress(98, 100, "Создание отчета...")
                report_file = self._create_report(
                    input_file,
                    output_folder,
                    data_processor.logger_obj if data_processor.logger_obj else ChangeLogger(),
                    mode,
                    self.processed_rows,
                    skipped_rows
                )

            # Финальный отчет
            if self.processed_rows > 0:
                success_msg = f"🎉 Конвертация успешно завершена! Создано 2 JSON файла"
                self._log_gui_and_console(success_msg, "SUCCESS", gui_callback)
                self._log_gui_and_console(f"📄 hunters.json: {len(hunters_data)} записей", "INFO", gui_callback)
                self._log_gui_and_console(f"📄 huntingtickets.json: {len(tickets_data)} записей", "INFO", gui_callback)
                if report_file:
                    self._log_gui_and_console(f"📊 Отчет создан: {report_file}", "INFO", gui_callback)
            else:
                self._log_gui_and_console("⚠️ Конвертация завершена, но данных нет!", "WARNING", gui_callback)

            self._update_progress(100, 100, "✅ Конвертация завершена", "00:00")

            return {
                'success': True,
                'processed_rows': self.processed_rows,
                'skipped_rows': skipped_rows,
                'output_folder': output_folder,
                'report_file': report_file,
                'log_file': self.log_file,
                'enriched_data': self.enriched_data
            }

        except Exception as e:
            error_msg = f"❌ Ошибка конвертации: {str(e)}\n{traceback.format_exc()}"
            self._log_gui_and_console(error_msg, "ERROR", gui_callback)
            self.logger.error(error_msg, exc_info=True)
            return {
                'success': False,
                'error': str(e)
            }

    def _format_hunter_to_json(self, row: Dict[str, Any], selected_region: str = None) -> Dict[str, Any]:
        """Преобразование строки в формат hunters.json (по эталону 6.1._Реестр_охотников__1_.json)"""
        result = {}

        # Функция для преобразования любого значения в строку
        def to_string(value):
            if pd.isna(value) or value == "":
                return ""
            # Все значения преобразуем в строку
            return str(value)

        # 1. date_entry
        date_entry = row.get('date_entry', '')
        result["date_entry"] = to_string(date_entry)

        # 2. municipality (объект)
        municipality = {}
        municipality_code = row.get('municipality_code', "")
        municipality_name = row.get('municipality_name', "")
        
        municipality["code"] = to_string(municipality_code)
        if municipality_code and municipality_name:
            municipality["name"] = to_string(f"{municipality_code} - {municipality_name}")
        else:
            municipality["name"] = to_string(municipality_name or municipality_code)

        result["municipality"] = municipality

        # 3. Основные поля (все в кавычках)
        for field in ['surname', 'hunter_name', 'patronymic', 'birth_date', 
                     'birth_place', 'postal_address', 'postal_code', 'phone', 
                     'email', 'address']:
            result[field] = to_string(row.get(field, ""))

        # 4. snils_code
        snils_code = row.get('snils_code', '')
        result["snils_code"] = to_string(snils_code)

        # 5. identity_type (объект)
        identity_type = {}
        identity_type_code = row.get('identity_type_code', "")
        identity_type_name = row.get('identity_type_name', "")

        identity_type["code"] = to_string(identity_type_code)
        identity_type["name"] = to_string(identity_type_name)

        result["identity_type"] = identity_type

        # 6. Паспортные данные (серия и номер - как строки без пробелов)
        for field in ['series_passport', 'number_passport', 'date_issue', 'issued_by']:
            value = row.get(field, "")
            # Для серии и номера паспорта убираем пробелы
            if field in ['series_passport', 'number_passport'] and value:
                value_str = str(value).strip().replace(" ", "")
                result[field] = to_string(value_str)
            else:
                result[field] = to_string(value)

        # 7. nationality (объект)
        nationality = {}
        nationality_code = row.get('nationality_code', "")
        nationality_name = row.get('nationality_name', "")

        nationality["code"] = to_string(nationality_code)
        nationality["name"] = to_string(nationality_name)

        result["nationality"] = nationality

        # 8. Остальные поля
        result["link"] = to_string(row.get('link', ""))
        result["traditional_residence_places"] = []

        # 9. organization_id (объект)
        organization_id = {}
        organization_type = row.get('organization_type', "")
        organization_inn = row.get('organization_inn', "")

        if organization_type:
            organization_id["organizations_type"] = {"name": to_string(organization_type)}
        if organization_inn:
            organization_id["unique_inn"] = to_string(organization_inn)

        result["organization_id"] = organization_id if organization_id else {}

        # 10. Охотничий билет (серия и номер - как строки без пробелов)
        for field in ['series_ticket', 'number_ticket', 'date_issue_ticket']:
            value = row.get(field, "")
            # Для серии и номера билета убираем пробелы
            if field in ['series_ticket', 'number_ticket'] and value:
                value_str = str(value).strip().replace(" ", "")
                result[field] = to_string(value_str)
            else:
                result[field] = to_string(value)

        # 11. Добавляем регион если выбран
        if selected_region:
            result["region_name"] = to_string(selected_region)
            for code, name in RUSSIAN_REGIONS.items():
                if name == selected_region:
                    result["region_code"] = to_string(code)
                    break

        return result

    def _format_ticket_to_json(self, row: Dict[str, Any], row_idx: int) -> Dict[str, Any]:
        """Преобразование строки в формат huntingtickets.json (по эталону huntingtickets.json)"""
        result = {}

        # Все значения должны быть в кавычках
        def to_string(value):
            if pd.isna(value) or value == "":
                return ""
            return str(value)

        # Основные поля
        for field in ['date_entry', 'series', 'number', 'date_issue']:
            value = ""
            if field == 'series':
                value = row.get('series_ticket', "")
                # Убираем пробелы из серии билета
                if value:
                    value = str(value).strip().replace(" ", "")
            elif field == 'number':
                value = row.get('number_ticket', "")
                # Убираем пробелы из номера билета
                if value:
                    value = str(value).strip().replace(" ", "")
            elif field == 'date_issue':
                value = row.get('date_issue_ticket', "")
            else:
                value = row.get(field, "")
            
            result[field] = to_string(value)

        # Данные охотника (объект)
        hunter_id = {}
        for field in ['series_passport', 'number_passport', 'date_issue', 'issued_by']:
            value = row.get(field, "")
            # Для серии и номера паспорта убираем пробелы
            if field in ['series_passport', 'number_passport'] and value:
                value = str(value).strip().replace(" ", "")
            hunter_id[field] = to_string(value)

        result["hunter_id"] = hunter_id

        # Коренные народы (строковое значение "true" или "false")
        indigenous = row.get('is_belonged_to_indigenous_people', '')
        if indigenous and str(indigenous).strip():
            indigenous_str = str(indigenous).strip().lower()
            if indigenous_str in ['true', 'false']:
                result["is_belonged_to_indigenous_people"] = to_string(indigenous_str)
            else:
                result["is_belonged_to_indigenous_people"] = to_string("false")
        else:
            result["is_belonged_to_indigenous_people"] = to_string("false")

        # cancellation_date
        cancellation_date = row.get('cancellation_date', '')
        result["cancellation_date"] = to_string(cancellation_date)

        # cancellation_reason (из справочника)
        cancellation_reason = {}
        cancellation_reason_code = row.get('cancellation_reason_code', '')
        
        if cancellation_reason_code and cancellation_reason_code in CANCELLATION_REASONS:
            reason_name = CANCELLATION_REASONS[cancellation_reason_code]
            cancellation_reason["name"] = to_string(reason_name)
        else:
            cancellation_reason["name"] = to_string("")

        result["cancellation_reason"] = cancellation_reason

        return result

    def _validate_date_format(self, date_str: str) -> bool:
        """Проверка формата даты ГГГГ-ММ-ДД"""
        if not date_str or str(date_str).strip() == "":
            return True
        
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        if not re.match(pattern, str(date_str).strip()):
            return False
        
        try:
            year, month, day = map(int, str(date_str).strip().split('-'))
            if month < 1 or month > 12:
                return False
            if day < 1 or day > 31:
                return False
            
            datetime(year, month, day)
            return True
        except:
            return False

    def _load_data(self, input_file: Path, sheet_name: str = None) -> pd.DataFrame:
        """Загрузка данных из Excel или CSV файла"""
        try:
            file_extension = input_file.suffix.lower()

            if file_extension in ['.xlsx', '.xlsm', '.xls']:
                df = self._load_excel_data(input_file, sheet_name)
            elif file_extension == '.csv':
                df = self._load_csv_data(input_file)
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {file_extension}")

            if self.logger:
                self.logger.info(f"Загружено {len(df)} строк, {len(df.columns)} столбцов")
                self.logger.info(f"Названия столбцов: {list(df.columns)}")

                missing_required = []
                for field in REQUIRED_FIELDS:
                    if field not in df.columns:
                        missing_required.append(field)

                if missing_required:
                    self.logger.warning(f"Отсутствуют обязательные поля: {missing_required}")

            return df

        except Exception as e:
            self.logger.error(f"Ошибка загрузки файла: {e}")
            raise

    def _load_csv_data(self, input_file: Path) -> pd.DataFrame:
        """Загрузка данных из CSV файла"""
        try:
            encodings = ['utf-8-sig', 'cp1251', 'windows-1251', 'utf-8']

            for encoding in encodings:
                try:
                    with open(input_file, 'r', encoding=encoding) as f:
                        sample = f.read(4096)
                        f.seek(0)

                        delimiter = ',' if ',' in sample else ';'

                        df = pd.read_csv(
                            input_file,
                            delimiter=delimiter,
                            encoding=encoding,
                            on_bad_lines='skip'
                        )
                    break
                except Exception as e:
                    if encoding == encodings[-1]:
                        raise e
                    continue

            df.columns = df.columns.astype(str).str.strip()
            df = df.dropna(axis=1, how='all')

            self.logger.info(f"Загружено строк из CSV: {len(df)}")
            return df

        except Exception as e:
            self.logger.error(f"Ошибка загрузки CSV файла: {e}")
            raise

    def _load_excel_data(self, input_file: Path, sheet_name: str = None) -> pd.DataFrame:
        """Загрузка данных из Excel файла"""
        try:
            workbook = load_workbook(input_file, read_only=True, data_only=True)

            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            data = list(sheet.values)

            if not data:
                workbook.close()
                return pd.DataFrame()

            # Обработка заголовков
            if len(data) >= 2:
                first_row = [str(cell) if cell is not None else "" for cell in data[0]]
                second_row = [str(cell) if cell is not None else "" for cell in data[1]]

                latin_count = sum(1 for cell in first_row if any('a' <= char.lower() <= 'z' for char in str(cell)))
                cyrillic_count = sum(1 for cell in second_row if any('\u0400' <= char <= '\u04FF' for char in str(cell)))

                if latin_count > 0 and cyrillic_count > 0:
                    columns = first_row
                    data_rows = data[2:]
                else:
                    columns = first_row
                    data_rows = data[1:]
            else:
                columns = [str(cell) if cell is not None else "" for cell in data[0]]
                data_rows = data[1:]

            df = pd.DataFrame(data_rows, columns=columns)
            workbook.close()

            df = df.dropna(axis=1, how='all')
            df.columns = df.columns.astype(str).str.strip()

            self.logger.info(f"Загружено строк из Excel: {len(df)}")
            return df

        except Exception as e:
            if 'workbook' in locals():
                workbook.close()
            self.logger.error(f"Ошибка загрузки Excel файла: {e}")
            raise

    def _save_json(self, data: List[Dict], output_file: Path):
        """Сохранение данных в JSON файл с правильным форматированием"""
        try:
            # Функция для правильного форматирования JSON
            def format_json_item(item):
                if isinstance(item, dict):
                    result = {}
                    for key, value in item.items():
                        # Обрабатываем вложенные структуры
                        if isinstance(value, dict):
                            result[key] = format_json_item(value)
                        elif isinstance(value, list):
                            result[key] = [format_json_item(v) for v in value]
                        elif isinstance(value, bool):
                            # Логические значения сохраняем как строки "true"/"false"
                            result[key] = "true" if value else "false"
                        elif value is None or (isinstance(value, str) and value == ""):
                            # Пустые значения
                            result[key] = ""
                        else:
                            # Все остальные значения как строки
                            result[key] = str(value)
                    return result
                elif isinstance(item, list):
                    return [format_json_item(i) for i in item]
                elif isinstance(item, bool):
                    return "true" if item else "false"
                else:
                    return str(item) if item is not None else ""

            # Форматируем все данные
            formatted_data = [format_json_item(item) for item in data]

            # Сохраняем с правильным форматированием
            with open(output_file, 'w', encoding='utf-8') as f:
                json_str = json.dumps(formatted_data, ensure_ascii=False, indent=2)
                f.write(json_str)
            self.logger.info(f"Данные сохранены в {output_file}")
        except Exception as e:
            self.logger.error(f"Ошибка сохранения JSON: {e}")
            raise

    def _create_report(self, input_file: Path, output_folder: Path,
                       change_logger: ChangeLogger, mode: str,
                       processed_rows: int, skipped_rows: int = 0) -> Path:
        """Создание отчета о конвертации"""
        try:
            report_file = input_file.parent / f"{input_file.stem}_conversion_report.txt"
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=" * 60 + "\n")
                f.write("ОТЧЕТ О КОНВЕРТАЦИИ EXCEL/CSV В JSON\n")
                f.write("=" * 60 + "\n")
                f.write(f"Дата создания: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Исходный файл: {input_file}\n")
                f.write(f"Выходная папка: {output_folder}\n")
                f.write(f"Режим конвертации: {mode}\n")
                f.write(f"Обработано строк: {processed_rows}\n")
                f.write(f"Пропущено строк: {skipped_rows}\n")
                f.write(f"Созданные файлы:\n")
                f.write(f"  - hunters.json ({processed_rows} записей)\n")
                f.write(f"  - huntingtickets.json ({processed_rows} записей)\n")

                summary = change_logger.get_summary()
                f.write(f"Изменения в данных: {summary['total_changes']}\n")
                f.write(f"Ошибки валидации: {summary['total_errors']}\n")
                f.write(f"Обогащенные поля: {summary['total_enrichments']}\n")
                f.write(f"Отсутствуют обязательные поля: {summary['total_missing_required']}\n")
                f.write(f"Пропущенные строки: {summary['total_skipped_rows']}\n")

                if summary['skipped_rows']:
                    f.write("\nПРОПУЩЕННЫЕ СТРОКИ:\n")
                    f.write("-" * 40 + "\n")
                    for skipped in summary['skipped_rows']:
                        f.write(f"Строка {skipped['row']}: {skipped['reason']}\n")
                    f.write("\n")

                if summary['missing_required']:
                    f.write("ОТСУТСТВУЮЩИЕ ОБЯЗАТЕЛЬНЫЕ ПОЛЯ:\n")
                    f.write("-" * 40 + "\n")
                    for row, fields in summary['missing_required'].items():
                        field_names = [field['field'] for field in fields]
                        f.write(f"Строка {row}: {', '.join(field_names)}\n")
                    f.write("\n")

                if summary['errors']:
                    f.write("ДЕТАЛИ ОШИБОК:\n")
                    f.write("-" * 40 + "\n")
                    for row, errors in summary['errors'].items():
                        for error in errors:
                            f.write(f"Строка {row}: {error['field']} = '{error['value']}' - {error['error']}\n")
                    f.write("\n")

                f.write("=" * 60 + "\n")
                f.write("КОНЕЦ ОТЧЕТА\n")
                f.write("=" * 60 + "\n")

            self.logger.info(f"Отчет создан: {report_file}")
            return report_file

        except Exception as e:
            self.logger.error(f"Ошибка создания отчета: {e}")
            return None


# ============================================================================
# ЗАПУСК ПРОГРАММЫ
# ============================================================================
if __name__ == "__main__":
    main()
